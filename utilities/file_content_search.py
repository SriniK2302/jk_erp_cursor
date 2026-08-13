"""Search files for an exact phrase within the first N words of extracted text.

Supports plain text, Excel (.xls / .xlsx-like), PDF (pypdf), Word .docx, and PowerPoint .pptx
(zip + XML text). Other suffixes are skipped.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal

_FIRST_WORDS = 500
_READ_CAP_BYTES = 2_000_000

_TEXT_SUFFIXES = frozenset(
    {
        ".txt",
        ".md",
        ".csv",
        ".tsv",
        ".json",
        ".xml",
        ".html",
        ".htm",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".jsx",
        ".mjs",
        ".cjs",
        ".css",
        ".sql",
        ".yaml",
        ".yml",
        ".ini",
        ".cfg",
        ".toml",
        ".rst",
        ".bat",
        ".ps1",
        ".sh",
        ".env",
    }
)

_XLSX_LIKE_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})

FileKind = Literal["text", "xlsx_like", "xls", "pdf", "docx", "pptx", "skip"]


@dataclass
class FileContentSearchReport:
    root: Path
    phrase: str
    word_limit: int
    scanned_files: int = 0
    skipped_files: int = 0
    matches: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def first_n_words_slice(text: str, n: int) -> str:
    """Return the prefix of `text` through the end of the n-th word (whitespace preserved)."""
    if n <= 0 or not text:
        return ""
    i = 0
    word_count = 0
    length = len(text)
    while i < length and word_count < n:
        while i < length and text[i].isspace():
            i += 1
        if i >= length:
            break
        word_count += 1
        while i < length and not text[i].isspace():
            i += 1
    return text[:i]


def _read_limited_bytes(path: Path) -> str:
    with path.open("rb") as handle:
        raw = handle.read(_READ_CAP_BYTES)
    return raw.decode("utf-8", errors="replace")


def classify_file_kind(path: Path) -> FileKind:
    suffix = path.suffix.lower()
    if suffix in _XLSX_LIKE_SUFFIXES:
        return "xlsx_like"
    if suffix == ".xls":
        return "xls"
    if suffix == ".pdf":
        return "pdf"
    if suffix == ".docx":
        return "docx"
    if suffix == ".pptx":
        return "pptx"
    if suffix in _TEXT_SUFFIXES:
        return "text"
    return "skip"


def phrase_in_first_words(text: str, phrase: str, word_limit: int) -> bool:
    if not phrase:
        return False
    window = first_n_words_slice(text, word_limit)
    return phrase.casefold() in window.casefold()


class _StreamingPhraseMatch:
    """Feed cell/string tokens; detect phrase within first `word_limit` words without scanning the whole sheet."""

    __slots__ = ("_phrase_cf", "_word_limit", "_text", "_max_len")

    def __init__(self, phrase_cf: str, word_limit: int, max_joined_len: int = 1_500_000) -> None:
        self._phrase_cf = phrase_cf
        self._word_limit = word_limit
        self._text = ""
        self._max_len = max_joined_len

    def feed(self, raw: str) -> None:
        t = raw.strip()
        if not t:
            return
        self._text = (self._text + " " + t) if self._text else t
        if len(self._text) > self._max_len:
            self._text = self._text[: self._max_len]

    def matched(self) -> bool:
        window = first_n_words_slice(self._text, self._word_limit)
        return self._phrase_cf in window.casefold()

    def exhausted_without_match(self) -> bool:
        """True once text extends past the first-`word_limit`-words window and the phrase is not there."""
        window = first_n_words_slice(self._text, self._word_limit)
        return len(self._text) > len(window) and self._phrase_cf not in window.casefold()

    @property
    def accumulated(self) -> str:
        return self._text


def _match_xlsx_like_file(path: Path, phrase: str, word_limit: int) -> bool:
    from openpyxl import load_workbook

    phrase_cf = phrase.strip().casefold()
    if not phrase_cf:
        return False

    matcher = _StreamingPhraseMatch(phrase_cf, word_limit)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        cell_budget = 400_000
        seen = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    seen += 1
                    if seen > cell_budget:
                        return matcher.matched()
                    if v is None:
                        continue
                    matcher.feed(str(v))
                    if matcher.matched():
                        return True
                    if matcher.exhausted_without_match():
                        return False
        return matcher.matched()
    finally:
        wb.close()


def _match_xls_file(path: Path, phrase: str, word_limit: int) -> bool:
    import xlrd

    phrase_cf = phrase.strip().casefold()
    if not phrase_cf:
        return False

    matcher = _StreamingPhraseMatch(phrase_cf, word_limit)
    book = xlrd.open_workbook(path, on_demand=True, formatting_info=False)
    try:
        row_cap = 8000
        col_cap = 256
        cell_budget = 400_000
        seen = 0
        for si in range(book.nsheets):
            sh = book.sheet_by_index(si)
            nrows = min(sh.nrows, row_cap)
            ncols = min(sh.ncols, col_cap)
            for ri in range(nrows):
                for ci in range(ncols):
                    seen += 1
                    if seen > cell_budget:
                        return matcher.matched()
                    v = sh.cell_value(ri, ci)
                    if v is None or v == "":
                        continue
                    if isinstance(v, float) and v == int(v):
                        s = str(int(v))
                    else:
                        s = str(v)
                    matcher.feed(s)
                    if matcher.matched():
                        return True
                    if matcher.exhausted_without_match():
                        return False
        return matcher.matched()
    finally:
        book.release_resources()


def _match_text_file(path: Path, phrase: str, word_limit: int) -> bool:
    try:
        text = _read_limited_bytes(path)
    except OSError:
        return False
    return phrase_in_first_words(text, phrase, word_limit)


def _extract_pdf_text(path: Path, *, max_pages: int = 30, max_chars: int = 240_000) -> str:
    from pypdf import PdfReader

    try:
        reader = PdfReader(path)
    except Exception:
        return ""
    parts: list[str] = []
    try:
        n = min(len(reader.pages), max_pages)
    except Exception:
        return ""
    for i in range(n):
        try:
            page = reader.pages[i]
            t = page.extract_text() or ""
        except Exception:
            t = ""
        parts.append(t)
    return "\n".join(parts)[:max_chars]


def _match_pdf_file(path: Path, phrase: str, word_limit: int) -> bool:
    text = _extract_pdf_text(path)
    if not text.strip():
        return False
    return phrase_in_first_words(text, phrase, word_limit)


def _extract_docx_text(path: Path, *, max_chars: int = 400_000) -> str:
    import zipfile
    from xml.etree import ElementTree as ET

    try:
        with zipfile.ZipFile(path) as zf:
            data = zf.read("word/document.xml")
    except (KeyError, OSError, zipfile.BadZipFile):
        return ""
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return ""
    texts: list[str] = []
    for el in root.iter():
        if el.text and el.text.strip():
            texts.append(el.text)
        if el.tail and el.tail.strip():
            texts.append(el.tail)
    return " ".join(texts)[:max_chars]


def _match_docx_file(path: Path, phrase: str, word_limit: int) -> bool:
    text = _extract_docx_text(path)
    if not text.strip():
        return False
    return phrase_in_first_words(text, phrase, word_limit)


def _extract_pptx_text(path: Path, *, max_chars: int = 400_000) -> str:
    import zipfile
    from xml.etree import ElementTree as ET

    texts: list[str] = []
    try:
        with zipfile.ZipFile(path) as zf:
            names = sorted(
                n
                for n in zf.namelist()
                if n.startswith("ppt/slides/slide") and n.endswith(".xml")
            )
            for name in names:
                try:
                    data = zf.read(name)
                    root = ET.fromstring(data)
                except (KeyError, OSError, ET.ParseError):
                    continue
                for el in root.iter():
                    if el.text and el.text.strip():
                        texts.append(el.text)
                    if el.tail and el.tail.strip():
                        texts.append(el.tail)
                if len(" ".join(texts)) >= max_chars:
                    break
    except (OSError, zipfile.BadZipFile):
        return ""
    return " ".join(texts)[:max_chars]


def _match_pptx_file(path: Path, phrase: str, word_limit: int) -> bool:
    text = _extract_pptx_text(path)
    if not text.strip():
        return False
    return phrase_in_first_words(text, phrase, word_limit)


def _file_matches(path: Path, kind: FileKind, phrase: str, word_limit: int) -> bool:
    if kind == "text":
        return _match_text_file(path, phrase, word_limit)
    if kind == "xlsx_like":
        return _match_xlsx_like_file(path, phrase, word_limit)
    if kind == "xls":
        return _match_xls_file(path, phrase, word_limit)
    if kind == "pdf":
        return _match_pdf_file(path, phrase, word_limit)
    if kind == "docx":
        return _match_docx_file(path, phrase, word_limit)
    if kind == "pptx":
        return _match_pptx_file(path, phrase, word_limit)
    return False


def scan_folder_for_phrase(
    root: Path,
    phrase: str,
    *,
    word_limit: int = _FIRST_WORDS,
) -> FileContentSearchReport:
    report = FileContentSearchReport(root=root, phrase=phrase, word_limit=word_limit)
    phrase_stripped = phrase.strip()
    if not phrase_stripped:
        return report

    try:
        resolved_root = root.expanduser().resolve()
    except OSError as exc:
        report.errors.append(f"Invalid folder: {exc}")
        return report

    if not resolved_root.is_dir():
        report.errors.append("Root is not a directory.")
        return report

    for path in resolved_root.rglob("*"):
        if not path.is_file():
            continue
        report.scanned_files += 1
        kind = classify_file_kind(path)
        if kind == "skip":
            report.skipped_files += 1
            continue
        try:
            if _file_matches(path, kind, phrase_stripped, word_limit):
                report.matches.append(str(path))
        except Exception as exc:
            report.errors.append(f"{path}: {exc}")

    return report
