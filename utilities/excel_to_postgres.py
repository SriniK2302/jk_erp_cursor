"""Import rows from an Excel sheet into PostgreSQL.

The target table must already exist (create it first, e.g. Data Utilities →
Create PostgreSQL table). Selected sheet columns are matched to table columns by
sanitized name (case-insensitive). Coercion uses each column’s type from
PostgreSQL. String values longer than VARCHAR/CHAR allow are truncated; TEXT is
unbounded.
"""

from __future__ import annotations

import csv
import math
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

ProgressCallback = Callable[[str, int, int | None, str], None]

import psycopg
from psycopg import Error as PsycopgError
from django.conf import settings
from openpyxl import load_workbook
from psycopg import sql
#from tkinter import Tk, TclError, filedialog

from utilities.pg_row_delete import connect_with_params

EXCEL_FILETYPES = [
    ("Excel and CSV files", "*.xlsx *.xlsm *.xltx *.xltm *.csv"),
    ("Excel workbooks", "*.xlsx *.xlsm *.xltx *.xltm"),
    ("CSV files", "*.csv"),
    ("All files", "*.*"),
]


# Synthetic single "sheet" name used for CSV files (which have no sheets).
CSV_SHEET_NAME = "Sheet1"

# Default width when information_schema has no length (e.g. VARCHAR without max)
MIN_VARCHAR_LEN = 255



def choose_excel_file() -> Path | None:
    try:
        from tkinter import Tk, TclError, filedialog
    except ImportError as exc:
        raise RuntimeError(
            "File picker requires a desktop environment with tkinter; "
            "not available on this server."
        ) from exc

    try:
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askopenfilename(
            title="Select Excel workbook",
            filetypes=EXCEL_FILETYPES,
        )
        root.destroy()
    except TclError as exc:
        raise RuntimeError("Unable to launch file picker.") from exc
    if not selected:
        return None
    return Path(selected)


def _csv_cell_value(raw: str | None) -> Any:
    """Infer int/float/bool from a CSV string cell; keep everything else as trimmed text."""
    if raw is None:
        return None
    s = raw.strip()
    if not s:
        return None
    if s.upper() in ("TRUE", "FALSE"):
        return s.upper() == "TRUE"
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def _read_csv_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return [], []
    headers = [_header_cell_str(c) for c in rows[0]]
    width = len(headers)
    data: list[list[Any]] = []
    for row in rows[1:]:
        if len(row) < width:
            row = row + [None] * (width - len(row))
        else:
            row = row[:width]
        data.append([_csv_cell_value(c) for c in row])
    return headers, data


def list_sheet_names(path: Path) -> list[str]:
    path = path.resolve()
    if path.suffix.lower() == ".csv":
        return [CSV_SHEET_NAME]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets]
    finally:
        wb.close()


def _locate_header_row(path: Path, sheet_name: str) -> tuple[int, int | None, int | None]:
    """
    Best-effort detection of where a sheet's real header row starts.

    1. If the sheet has a formal Excel Table (Insert > Table) defined,
       use its header row and column bounds.
    2. Otherwise, skip any leading fully-blank rows and use the first row
       that has at least one non-empty cell as the header.

    Returns (header_row, min_col, max_col), all 1-based/inclusive.
    min_col/max_col are ``None`` when there's no table to bound columns
    (the caller then uses each row's natural width).
    """
    if path.suffix.lower() == ".csv":
        return 1, None, None

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name]
        tables = getattr(ws, "tables", None)
        if tables:
            for name in tables:
                try:
                    from openpyxl.utils.cell import range_boundaries

                    min_col, min_row, max_col, _max_row = range_boundaries(tables[name].ref)
                    return min_row, min_col, max_col
                except Exception:
                    continue
    finally:
        wb.close()

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(c is not None and str(c).strip() != "" for c in row):
                return i, None, None
            if i >= 50:
                break
    finally:
        wb.close()
    return 1, None, None


def read_sheet_headers_only(path: Path, sheet_name: str) -> list[str]:
    path = path.resolve()
    if path.suffix.lower() == ".csv":
        headers, _ = _read_csv_rows(path)
        return headers
    header_row, min_col, max_col = _locate_header_row(path, sheet_name)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        first = next(
            ws.iter_rows(
                min_row=header_row, max_row=header_row,
                min_col=min_col, max_col=max_col, values_only=True,
            ),
            None,
        )
        if not first:
            return []
        return [_header_cell_str(c) for c in first]
    finally:
        wb.close()


def _header_cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _cell_raw(value: Any) -> Any:
    """Normalize openpyxl cell to Python value for typing (bool before int)."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return value
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value).strip() or None


def read_sheet_rows_raw(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    """Header row as strings; data rows preserve int/float/bool/datetime/str."""
    path = path.resolve()
    if path.suffix.lower() == ".csv":
        return _read_csv_rows(path)
    header_row, min_col, max_col = _locate_header_row(path, sheet_name)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(
            min_row=header_row, min_col=min_col, max_col=max_col, values_only=True
        )
        first = next(rows_iter, None)
        if first is None:
            return [], []
        headers = [_header_cell_str(c) for c in first]
        width = len(headers)
        data: list[list[Any]] = []
        for row in rows_iter:
            cells = list(row)
            if len(cells) < width:
                cells = cells + [None] * (width - len(cells))
            else:
                cells = cells[:width]
            data.append([_cell_raw(c) for c in cells])
        return headers, data
    finally:
        wb.close()
        


def read_sheet_rows(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    """First row = headers. Data rows as strings (legacy helper)."""
    headers, raw = read_sheet_rows_raw(path, sheet_name)
    str_rows = [[_cell_str_from_raw(c) for c in row] for row in raw]
    return headers, str_rows


def _cell_str_from_raw(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def sanitize_identifier(name: str, max_len: int = 63) -> str:
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", (name or "").strip()).lower()
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return "col"
    if s[0].isdigit():
        s = "c_" + s
    return s[:max_len]


def unique_column_names(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in headers:
        base = sanitize_identifier(raw)
        cnt = seen.get(base, 0)
        seen[base] = cnt + 1
        if cnt == 0:
            out.append(base)
        else:
            out.append(f"{base}_{cnt}")
    return out


def _selected_header_indices(
    headers: list[str], selected_headers: list[str] | None
) -> list[int]:
    """Indices into ``headers`` for import, in sheet order. Empty selection = all columns."""
    if not selected_headers:
        return list(range(len(headers)))
    allowed = set(headers)
    bad = [h for h in selected_headers if h not in allowed]
    if bad:
        raise ValueError(
            "Unknown column(s) in selection: " + ", ".join(repr(b) for b in bad[:5])
            + ("…" if len(bad) > 5 else "")
        )
    picked = set(selected_headers)
    return [i for i, h in enumerate(headers) if h in picked]


def _existing_public_table_columns(cur, table_name: str) -> list[str] | None:
    """Return column attnames in order if ``public.table_name`` exists, else None."""
    cur.execute(
        """
        SELECT c.oid
        FROM pg_catalog.pg_class c
        INNER JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = 'public'
          AND c.relkind IN ('r', 'p')
          AND pg_catalog.lower(c.relname) = pg_catalog.lower(%s)
        LIMIT 1
        """,
        (table_name,),
    )
    row = cur.fetchone()
    if not row:
        return None
    oid = row[0]
    cur.execute(
        """
        SELECT a.attname::text
        FROM pg_catalog.pg_attribute a
        WHERE a.attrelid = %s
          AND a.attnum > 0
          AND NOT a.attisdropped
        ORDER BY a.attnum
        """,
        (oid,),
    )
    return [r[0] for r in cur.fetchall()]


def _fetch_public_column_metadata(
    cur, table_name: str
) -> dict[str, dict[str, Any]]:
    """
    lower(column_name) -> column metadata for ``public`` table (case-insensitive name).

    Includes fields needed for coercion without scanning sheet rows for types,
    plus identity/serial/generated flags used to exclude auto and calculated
    columns from Excel-to-table matching.
    """
    cur.execute(
        """
        SELECT column_name, data_type, character_maximum_length,
               numeric_precision, numeric_scale, udt_name,
               is_identity, column_default, is_generated
        FROM information_schema.columns
        WHERE table_catalog = current_database()
          AND table_schema = 'public'
          AND lower(table_name) = lower(%s)
        """,
        (table_name,),
    )
    out: dict[str, dict[str, Any]] = {}
    for row in cur.fetchall():
        cname, dtype, cmax, nprec, nscale, udt, is_identity, col_default, is_generated = row
        is_auto = (is_identity == "YES") or bool(
            col_default and "nextval(" in col_default
        )
        is_calculated = is_generated == "ALWAYS"
        out[cname.lower()] = {
            "column_name": cname,
            "data_type": dtype,
            "character_maximum_length": cmax,
            "numeric_precision": nprec,
            "numeric_scale": nscale,
            "udt_name": udt,
            "is_auto": is_auto,
            "is_calculated": is_calculated,
        }
    return out



def _coercion_pg_type_from_db_meta(meta: dict[str, Any] | None) -> str:
    """
    Build a type string compatible with ``_coerce_for_pg`` / ``_effective_coerce_pg_type``
    from ``information_schema.columns`` metadata (no sheet scan).
    """
    if not meta:
        return "TEXT"
    dt = (meta.get("data_type") or "").lower()
    udt = (meta.get("udt_name") or "").lower()
    cmax = meta.get("character_maximum_length")
    nprec = meta.get("numeric_precision")
    nscale = meta.get("numeric_scale")

    if dt == "boolean":
        return "BOOLEAN"
    if dt in ("integer", "smallint"):
        return "INTEGER"
    if dt == "bigint":
        return "BIGINT"
    if dt == "double precision":
        return "DOUBLE PRECISION"
    if dt == "real":
        return "DOUBLE PRECISION"
    if dt in ("numeric", "decimal"):
        p = int(nprec) if nprec is not None else 18
        s = int(nscale) if nscale is not None else 6
        p = max(1, min(p, 1000))
        s = max(0, min(s, 1000))
        return f"NUMERIC({p},{s})"
    if dt == "date":
        return "DATE"
    if "timestamp" in dt:
        if "time zone" in dt or dt == "timestamp with time zone":
            return "TIMESTAMP WITH TIME ZONE"
        return "TIMESTAMP WITHOUT TIME ZONE"
    if dt == "text":
        return "TEXT"
    if dt == "character varying":
        w = int(cmax) if cmax is not None else MIN_VARCHAR_LEN
        return f"VARCHAR({max(w, MIN_VARCHAR_LEN)})"
    if dt == "character":
        w = int(cmax) if cmax is not None else 1
        return f"CHAR({max(w, 1)})"
    if dt == "uuid":
        return "UUID"
    if dt in ("json", "jsonb"):
        return "TEXT"
    if dt == "USER-DEFINED" and udt:
        # enums and other UDTs: coerce as text
        return "TEXT"
    return "TEXT"


def _infer_string_kind_width(inferred_pg_type: str) -> tuple[str | None, int | None]:
    """Return ('text', None), ('varchar', width), or (None, None) if not a string type."""
    t = inferred_pg_type.strip()
    u = t.upper()
    if u == "TEXT":
        return ("text", None)
    m = re.match(r"VARCHAR\s*\(\s*(\d+)\s*\)", t, re.I)
    if m:
        return ("varchar", int(m.group(1)))
    return (None, None)


def _string_max_chars_for_insert(
    pg_type: str, db_column_meta: dict[str, Any] | None
) -> int | None:
    """
    Maximum string length allowed for INSERT, or None if unbounded (TEXT).
    Prefer live DB metadata when appending to an existing table.
    """
    if db_column_meta:
        dt = (db_column_meta.get("data_type") or "").lower()
        if dt == "text":
            return None
        cmax = db_column_meta.get("character_maximum_length")
        if cmax is not None:
            return int(cmax)
    kind, width = _infer_string_kind_width(pg_type)
    if kind == "text":
        return None
    if kind == "varchar" and width is not None:
        return width
    return None


def _match_import_names_to_db(
    import_pg_names: list[str], existing_cols: list[str]
) -> tuple[list[tuple[int, str]], list[str]]:
    """
    Map subset column index -> actual DB column name (case-insensitive match).
    Returns (matched pairs, unmatched sanitized names).
    """
    by_lower: dict[str, str] = {}
    for c in existing_cols:
        by_lower.setdefault(c.lower(), c)
    matched: list[tuple[int, str]] = []
    unmatched: list[str] = []
    for j, name in enumerate(import_pg_names):
        dbn = by_lower.get(name.lower())
        if dbn is not None:
            matched.append((j, dbn))
        else:
            unmatched.append(name)
    return matched, unmatched


def _validate_db_name(name: str) -> bool:
    return bool(re.fullmatch(r"[a-zA-Z_][a-zA-Z0-9_]{0,62}", name))


def _validate_table_name(name: str) -> bool:
    return bool(re.fullmatch(r"[a-zA-Z_][a-zA-Z0-9_]{0,62}", name))


def _header_index_map(headers: list[str]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers)}


def cell_display_for_filter(value: Any) -> str:
    """String form for comparing to user-typed filter values."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and not math.isnan(value) and value == int(value):
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(value).rstrip("0").rstrip(".") if "." in str(value) else str(value)
    return str(value).strip()


def row_passes_filters(
    row: list[Any],
    headers: list[str],
    filters: list[tuple[str, str]],
) -> bool:
    idx_map = _header_index_map(headers)
    for col_name, want in filters:
        col_name = (col_name or "").strip()
        want = (want or "").strip()
        if not col_name or not want:
            continue
        if col_name not in idx_map:
            continue
        i = idx_map[col_name]
        got = cell_display_for_filter(row[i] if i < len(row) else None)
        if got.strip() != want.strip():
            return False
    return True


def _parse_bool_str(s: str) -> bool | None:
    t = s.strip().lower()
    if t in ("true", "t", "yes", "y", "1"):
        return True
    if t in ("false", "f", "no", "n", "0"):
        return False
    return None


def _try_decimal(s: str) -> Decimal | None:
    try:
        return Decimal(s.strip().replace(",", ""))
    except (InvalidOperation, ValueError, AttributeError):
        return None


def _parse_flexible_timestamp_str(raw: str) -> datetime | None:
    """
    Parse common Excel / human date strings into naive datetimes (start of day
    or first instant of month when no day is given).
    """
    s = raw.strip()
    if not s:
        return None
    try:
        iso = s.replace("Z", "+00:00")
        base = iso.split("+")[0].strip()
        return datetime.fromisoformat(base)
    except ValueError:
        pass

    for fmt in (
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    # Compact YYMM (4 digits) or YYYYMM (6 digits); MM must be 01–12.
    if re.fullmatch(r"\d{4}", s) is not None:
        yy, mm = int(s[:2]), int(s[2:])
        if 1 <= mm <= 12:
            try:
                return datetime(2000 + yy, mm, 1)
            except ValueError:
                pass
    elif re.fullmatch(r"\d{6}", s) is not None:
        yyyymm_dt: datetime | None = None
        yyyy, mm = int(s[:4]), int(s[4:6])
        if 1800 <= yyyy <= 2199 and 1 <= mm <= 12:
            try:
                yyyymm_dt = datetime(yyyy, mm, 1)
            except ValueError:
                yyyymm_dt = None
        if yyyymm_dt is not None:
            return yyyymm_dt
        yy, mm, dd = int(s[:2]), int(s[2:4]), int(s[4:6])
        if 1 <= mm <= 12:
            try:
                return datetime(2000 + yy, mm, dd)
            except ValueError:
                pass

    # "MARCH 2022", "Mar 2022", "March-2024"
    m = re.match(r"^([A-Za-z]+)[\s,-]+(\d{4})\s*$", s)
    if m:
        mon_tok, yr_s = m.group(1), m.group(2)
        yr = int(yr_s)
        for mt in (mon_tok.title(), mon_tok[:3].title()):
            for fmt in ("%B %Y", "%b %Y"):
                try:
                    return datetime.strptime(f"{mt} {yr}", fmt)
                except ValueError:
                    continue
    return None


def _effective_coerce_pg_type(
    inferred: str, db_column_meta: dict[str, Any] | None
) -> str:
    """Prefer DB column type so TIMESTAMP/DATE columns get correct coercion."""
    if not db_column_meta:
        return inferred
    dt = (db_column_meta.get("data_type") or "").lower()
    if "timestamp" in dt:
        return "TIMESTAMP WITHOUT TIME ZONE"
    if dt == "date":
        return "DATE"
    return inferred


def _format_import_psycopg_error(
    exc: PsycopgError,
    *,
    sheet_name: str,
    sheet_row_1based: int,
    sel_headers: list[str],
    insert_plan: list[tuple[int, str]],
    payload: tuple[Any, ...] | None = None,
) -> str:
    """
    Add sheet row and column context. insert_plan maps subset index -> DB column name
    (only columns selected for import).
    """
    pg_col: str | None = None
    if exc.diag is not None:
        pg_col = exc.diag.column_name

    sheet_header: str | None = None
    if pg_col:
        for j, dbn in insert_plan:
            if dbn.lower() == pg_col.lower():
                sheet_header = sel_headers[j]
                break

    exc_msg = str(exc)
    param_idx: int | None = None
    if not pg_col:
        m = re.search(r"parameter\s+\$(\d+)", exc_msg, re.I)
        if m:
            pn = int(m.group(1))
            pi = pn - 1
            if 0 <= pi < len(insert_plan):
                j, dbn = insert_plan[pi]
                sheet_header = sel_headers[j]
                pg_col = dbn
                param_idx = pn

    lines: list[str] = [
        f'Sheet {sheet_name!r}, data row {sheet_row_1based} (row 1 is the header).',
        "Only columns you selected for import are written.",
    ]

    # PostgreSQL doesn't report a column name for plain "value too long"
    # errors, so find the actual offending value ourselves by comparing
    # each value in this row against the column's VARCHAR/CHAR limit.
    if pg_col is None and payload is not None:
        m = re.search(r"character varying\((\d+)\)|character\((\d+)\)", exc_msg, re.I)
        if m:
            limit = int(m.group(1) or m.group(2))
            for (j, dbn), value in zip(insert_plan, payload):
                if isinstance(value, str) and len(value) > limit:
                    lines.append(
                        f"Failing value: Excel column {sel_headers[j]!r} → PostgreSQL "
                        f"{dbn!r} (limit {limit} chars). Value ({len(value)} chars): "
                        f"{value!r}."
                    )
                    lines.append(str(exc).strip())
                    return "\n".join(lines)

    if sheet_header is not None and pg_col is not None:
        hint = (
            f'Failing column: Excel/header {sheet_header!r} → PostgreSQL {pg_col!r}.'
        )
        if param_idx is not None:
            hint += f" (matches server message parameter ${param_idx}.)"
        lines.append(hint)
    elif pg_col is not None:
        lines.append(f"Failing PostgreSQL column (from your selection): {pg_col!r}.")
    else:
        pairs = [f"{sel_headers[j]!r} → {db!r}" for j, db in insert_plan]
        lines.append("Selected import mapping (Excel → DB): " + "; ".join(pairs) + ".")

    lines.append(str(exc).strip())
    return "\n".join(lines)


def _coerce_for_pg(
    value: Any,
    pg_type: str,
    db_column_meta: dict[str, Any] | None = None,
) -> Any:
    if value is None or value == "":
        return None
    base = pg_type.upper().split("(")[0].strip()

    if "BOOLEAN" in pg_type or base == "BOOLEAN":
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            b = _parse_bool_str(value)
            if b is not None:
                return b
        return None

    if base in ("INTEGER", "INT"):
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str) and value.strip().lstrip("-").isdigit():
            return int(value.strip())
        return None

    if base == "BIGINT":
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            try:
                return int(value.strip().replace(",", ""))
            except ValueError:
                return None
        return None

    if "DOUBLE" in pg_type or base == "DOUBLE":
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", ""))
            except ValueError:
                return None
        return None

    if base == "NUMERIC" or base.startswith("NUMERIC"):
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            d = _try_decimal(value)
            return d
        return None

    if base == "DATE":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            st = value.strip()
            try:
                return date.fromisoformat(st[:10])
            except ValueError:
                pass
            dtp = _parse_flexible_timestamp_str(st)
            if dtp is not None:
                return dtp.date()
        return None

    if "TIMESTAMP" in pg_type:
        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value
        if isinstance(value, date) and not isinstance(value, datetime):
            return datetime.combine(value, time.min)
        if isinstance(value, str):
            st = value.strip()
            try:
                return datetime.fromisoformat(st.replace("Z", "+00:00").split("+")[0])
            except ValueError:
                pass
            return _parse_flexible_timestamp_str(st)
        return None

    # VARCHAR / TEXT / CHAR — truncate to column max length so INSERT never fails on length
    s = cell_display_for_filter(value) if not isinstance(value, str) else value
    if not isinstance(s, str):
        s = str(s)
    max_ch = _string_max_chars_for_insert(pg_type, db_column_meta)
    if max_ch is not None and len(s) > max_ch:
        s = s[:max_ch]
    return s


def _shorten_bank_account_name(raw: str) -> str:
    """
    'HDFC SB 00581000000873' -> 'HDFC SB 000873'.

    Bank name + account type + last 6 digits of the account number. This is
    the naming convention already used for accounts in source_bank_cash_acs
    (max 15 chars). "CC-04" style type suffixes are dropped (only "CC" is
    kept); the account-index digit isn't part of the actual account number.
    """
    tokens = (raw or "").strip().split()
    if len(tokens) < 2:
        return (raw or "").strip()[:15]
    bank = tokens[0]
    acc_type = tokens[1].split("-")[0]
    digits = "".join(re.findall(r"\d", " ".join(tokens[2:])))
    last6 = digits[-6:] if len(digits) >= 6 else digits
    return f"{bank} {acc_type} {last6}".strip()[:15]


def _ensure_source_bank_cash_accounts(
    cur,
    *,
    insert_plan: list[tuple[int, str]],
    sel_idx: list[int],
    headers: list[str],
    data_rows: list[list[Any]],
) -> list[str]:
    """
    For an import into ``bank_transactions_source``: scan the sheet's
    ``source_ac`` values, and insert any that don't yet exist in
    ``source_bank_cash_acs`` (bank_name/account_type filled in from the
    sheet where available, fb_code left blank). Returns the list of newly
    created account codes.
    """
    source_ac_j = next((j for j, dbn in insert_plan if dbn.lower() == "source_ac"), None)
    if source_ac_j is None:
        return []
    account_type_j = next((j for j, dbn in insert_plan if dbn.lower() == "account_type"), None)

    seen: dict[str, str] = {}
    for full_row in data_rows:
        row = full_row
        if len(row) < len(headers):
            row = list(row) + [None] * (len(headers) - len(row))
        val = row[sel_idx[source_ac_j]]
        if val is None:
            continue
        sac = str(val).strip()
        if not sac:
            continue
        sac = _shorten_bank_account_name(sac)
        if sac in seen:
            continue
        atype = ""
        if account_type_j is not None:
            atv = row[sel_idx[account_type_j]]
            atype = str(atv).strip()[:2] if atv else ""
        seen[sac] = atype

    if not seen:
        return []

    too_long = [sac for sac in seen if len(sac) > 15]
    if too_long:
        raise ValueError(
            "These Source Ac values are longer than 15 characters (the limit "
            "for the accounts master table) and can't be auto-created: "
            + ", ".join(f"{sac!r} ({len(sac)} chars)" for sac in too_long)
            + ". Shorten them in the sheet before importing."
        )

    cur.execute(
        "SELECT source_ac FROM source_bank_cash_acs WHERE source_ac = ANY(%s)",
        (list(seen.keys()),),
    )

    existing = {row[0] for row in cur.fetchall()}
    to_create = [sac for sac in seen if sac not in existing]

    for sac in to_create:
        bank_name = sac.split(" ", 1)[0][:50]
        cur.execute(
            "INSERT INTO source_bank_cash_acs "
            "(source_ac, bank_name, account_type, fb_code) "
            "VALUES (%s, %s, %s, %s)",
            (sac, bank_name, seen[sac], ""),
        )

    return to_create


def import_sheet_to_postgres(
    file_path: Path,
    sheet_name: str,
    postgres_db: str,
    table_name: str,
    filter_triples: list[tuple[str, str]] | None = None,
    progress_callback: ProgressCallback | None = None,
    selected_headers: list[str] | None = None,
) -> dict[str, Any]:
    """
    Import selected sheet columns into existing ``table_name``. Column types and
    coercion come from PostgreSQL metadata (not from the sheet). Headers are
    sanitized to match table column names (case-insensitive).
    """
    if not _validate_db_name(postgres_db):
        raise ValueError(
            "Invalid database name. Use letters, digits, underscore; start with letter or underscore."
        )
    if not _validate_table_name(table_name):
        raise ValueError(
            "Invalid table name. Use letters, digits, underscore; start with letter or underscore."
        )

    file_path = file_path.resolve()
    if not file_path.is_file():
        raise ValueError(f"File not found: {file_path}")

    if progress_callback:
        progress_callback("reading", 0, None, "Reading workbook…")

    headers, data_rows = read_sheet_rows_raw(file_path, sheet_name)
    if not headers:
        raise ValueError("Sheet has no header row or is empty.")

    sel_idx = _selected_header_indices(headers, selected_headers)
    if not sel_idx:
        raise ValueError("Select at least one column to import.")

    all_pg_names = unique_column_names(headers)
    import_pg_names = [all_pg_names[i] for i in sel_idx]
    sel_headers = [headers[i] for i in sel_idx]

    nrows = len(data_rows)
    n_import = len(sel_idx)
    ft = filter_triples or []
    filters = [(a.strip(), b.strip()) for a, b in ft if a and b]

    inserted = 0
    db_cfg = settings.DATABASES["default"]
    conn = psycopg.connect(
        host=db_cfg.get("HOST") or "127.0.0.1",
        port=int(db_cfg.get("PORT") or 5432),
        user=db_cfg.get("USER") or "postgres",
        password=db_cfg.get("PASSWORD") or "",
        dbname=postgres_db,
        connect_timeout=int(db_cfg.get("OPTIONS", {}).get("connect_timeout", 10)),
    )
    unmatched_db: list[str] = []
    insert_db_names: list[str] = []
    auto_created_accounts: list[str] = []

    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            existing_cols = _existing_public_table_columns(cur, table_name)
            if existing_cols is None:
                raise ValueError(
                    f"Table {table_name!r} does not exist. Create it first "
                    "(Data Utilities → Create PostgreSQL table), then import."
                )

            matched, unmatched_db = _match_import_names_to_db(
                import_pg_names, existing_cols
            )
            if not matched:
                raise ValueError(
                    "None of the selected columns match the table’s column names "
                    "(case-insensitive). Check the table definition and your selection."
                )

            column_meta = _fetch_public_column_metadata(cur, table_name)

            # Database-calculated (generated) columns can't be written to directly;
            # skip them automatically rather than failing the whole import.
            skipped_calculated: list[str] = []
            insert_plan: list[tuple[int, str]] = []
            for j, dbn in matched:
                meta = column_meta.get(dbn.lower()) if column_meta else None
                if meta and meta.get("is_calculated"):
                    skipped_calculated.append(dbn)
                    continue
                insert_plan.append((j, dbn))

            if not insert_plan:
                raise ValueError(
                    "All matched columns are database-calculated (generated) columns "
                    "and can't be imported directly: " + ", ".join(skipped_calculated)
                    + ". Remove them from your selection."
                )

            insert_db_names = [db for _, db in insert_plan]
            col_by_lower = {c.lower(): c for c in existing_cols}
            pg_types: list[str] = []
            for j in range(n_import):
                dbn = col_by_lower.get(import_pg_names[j].lower())
                if dbn is not None:
                    meta = column_meta.get(dbn.lower()) if column_meta else None
                    pg_types.append(_coercion_pg_type_from_db_meta(meta))
                else:
                    pg_types.append("TEXT")

            # bank_transactions_source.source_ac always references the
            # source_bank_cash_acs master; auto-create any account the sheet
            # references but that doesn't exist yet, so historical imports
            # don't need manual account setup first.
            if table_name.lower() == "bank_transactions_source":
                auto_created_accounts = _ensure_source_bank_cash_accounts(
                    cur,
                    insert_plan=insert_plan,
                    sel_idx=sel_idx,
                    headers=headers,
                    data_rows=data_rows,
                )

            idents = [sql.Identifier(db) for _, db in insert_plan]
            placeholders = sql.SQL(", ").join(sql.Placeholder() for _ in insert_plan)
            insert_stmt = sql.SQL("INSERT INTO {} ({}) VALUES ({})").format(
                sql.Identifier(table_name),
                sql.SQL(", ").join(idents),
                placeholders,
            )

            shorten_accounts = table_name.lower() == "bank_transactions_source"
            narration_j = None
            opening_balance_skipped = 0
            if shorten_accounts:
                narration_j = next(
                    (j for j, dbn in insert_plan if dbn.lower() == "narration"), None
                )

            notify_every = max(1, nrows // 500) if nrows > 500 else 1
            for i, full_row in enumerate(data_rows, start=1):
                if progress_callback and (
                    i == 1 or i == nrows or i % notify_every == 0
                ):
                    progress_callback(
                        "inserting",
                        i,
                        nrows,
                        f"Writing to database… row {i:,} of {nrows:,}",
                    )
                row = full_row
                if len(row) < len(headers):
                    row = list(row) + [None] * (len(headers) - len(row))
                else:
                    row = row[: len(headers)]
                if not row_passes_filters(row, headers, filters):
                    continue
                sub_row = [row[j] for j in sel_idx]
                if narration_j is not None:
                    narration_val = sub_row[narration_j]
                    if narration_val and str(narration_val).strip().lower().startswith("opening balance"):
                        opening_balance_skipped += 1
                        continue
                if shorten_accounts:
                    for j, dbn in insert_plan:
                        if dbn.lower() == "source_ac" and sub_row[j] not in (None, ""):
                            sub_row[j] = _shorten_bank_account_name(str(sub_row[j]))
                payload = tuple(
                    _coerce_for_pg(
                        sub_row[j],
                        _effective_coerce_pg_type(
                            pg_types[j],
                            column_meta.get(dbn.lower()),
                        ),
                        column_meta.get(dbn.lower()),
                    )
                    for j, dbn in insert_plan
                )
                try:
                    cur.execute(insert_stmt, payload)
                except PsycopgError as e:
                    raise ValueError(
                        _format_import_psycopg_error(
                            e,
                            sheet_name=sheet_name,
                            sheet_row_1based=i,
                            sel_headers=sel_headers,
                            insert_plan=insert_plan,
                            payload=payload,
                        )
                    ) from e
                
                inserted += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    out: dict[str, Any] = {
        "inserted": inserted,
        "headers": sel_headers,
        "table_name": table_name,
        "database": postgres_db,
        "column_types": list(
            zip(
                insert_db_names,
                [pg_types[j] for j, _ in insert_plan],
                strict=True,
            )
        ),
        "table_existed": True,
        "import_sanitized_names": insert_db_names,
    }
    if unmatched_db:
        out["columns_not_in_table"] = unmatched_db
    if skipped_calculated:
        out["columns_calculated_skipped"] = skipped_calculated
    if auto_created_accounts:
        out["auto_created_accounts"] = auto_created_accounts
    if opening_balance_skipped:
        out["opening_balance_rows_skipped"] = opening_balance_skipped
    return out

def inspect_import_column_mapping(
    file_path: Path,
    sheet_name: str,
    postgres_db: str,
    table_name: str,
    selected_headers: list[str] | None = None,
) -> dict[str, Any]:
    """
    Compare selected Excel headers (sanitized for import) with existing table columns.

    Returns details used to warn users before import starts when names are not an
    exact match. Matching is case-insensitive against ``public.table_name``.
    """
    if not _validate_db_name(postgres_db):
        raise ValueError(
            "Invalid database name. Use letters, digits, underscore; start with letter or underscore."
        )
    if not _validate_table_name(table_name):
        raise ValueError(
            "Invalid table name. Use letters, digits, underscore; start with letter or underscore."
        )

    file_path = file_path.resolve()
    if not file_path.is_file():
        raise ValueError(f"File not found: {file_path}")

    headers = read_sheet_headers_only(file_path, sheet_name)
    if not headers:
        raise ValueError("Sheet has no header row or is empty.")

    sel_idx = _selected_header_indices(headers, selected_headers)
    if not sel_idx:
        raise ValueError("Select at least one column to import.")

    all_pg_names = unique_column_names(headers)
    selected_pairs = [
        {"excel_header": headers[i], "sanitized_name": all_pg_names[i]} for i in sel_idx
    ]
    selected_import_names = [p["sanitized_name"] for p in selected_pairs]
    selected_lower = {n.lower() for n in selected_import_names}

    db_cfg = settings.DATABASES["default"]
    column_meta: dict[str, dict[str, Any]] = {}
    conn = psycopg.connect(
        host=db_cfg.get("HOST") or "127.0.0.1",
        port=int(db_cfg.get("PORT") or 5432),
        user=db_cfg.get("USER") or "postgres",
        password=db_cfg.get("PASSWORD") or "",
        dbname=postgres_db,
        connect_timeout=int(db_cfg.get("OPTIONS", {}).get("connect_timeout", 10)),
    )
    try:
        with conn.cursor() as cur:
            existing_cols = _existing_public_table_columns(cur, table_name)
            if existing_cols is None:
                raise ValueError(
                    f"Table {table_name!r} does not exist. Create it first "
                    "(Data Utilities → Create PostgreSQL table), then import."
                )
            column_meta = _fetch_public_column_metadata(cur, table_name)
    finally:
        conn.close()

    by_lower = {c.lower(): c for c in existing_cols}
    missing_in_table = [
        name for name in selected_import_names if name.lower() not in by_lower
    ]
    matched_in_table = [
        by_lower[name.lower()]
        for name in selected_import_names
        if name.lower() in by_lower
    ]

    # Table columns not present in the selected Excel columns are expected
    # (auto/identity columns, calculated columns, or simply table columns
    # the source doesn't populate) and are never reported as a problem.
    table_only_columns = [
        c
        for c in existing_cols
        if c.lower() not in selected_lower
        and not (column_meta.get(c.lower()) or {}).get("is_auto")
        and not (column_meta.get(c.lower()) or {}).get("is_calculated")
    ]

    for pair in selected_pairs:
        matches = pair["sanitized_name"].lower() in by_lower
        pair["matches_table"] = matches
        meta = column_meta.get(pair["sanitized_name"].lower()) if matches else None
        if meta:
            pair["postgres_column"] = meta["column_name"]
            pair["postgres_type"] = _coercion_pg_type_from_db_meta(meta)
            pair["is_calculated"] = bool(meta.get("is_calculated"))
        else:
            pair["postgres_column"] = None
            pair["postgres_type"] = None
            pair["is_calculated"] = False

    exact_match = not missing_in_table
    return {
        "table_name": table_name,
        "selected_excel_count": len(selected_pairs),
        "table_column_count": len(existing_cols),
        "selected_columns": selected_pairs,
        "matched_columns": matched_in_table,
        "missing_in_table": missing_in_table,
        "table_only_columns": table_only_columns,
        "exact_match": exact_match,
    }



def _header_key_label(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def normalize_pg_type_for_ddl(raw: str) -> str:
    """
    Validate a PostgreSQL type string for DDL. Only a fixed whitelist / patterns
    are allowed (prevents SQL injection via the type column).
    """
    s = " ".join((raw or "").split()).strip()
    if not s:
        raise ValueError("Empty data_type cell.")
    u = s.upper()

    fixed: dict[str, str] = {
        "BOOLEAN": "BOOLEAN",
        "BOOL": "BOOLEAN",
        "INTEGER": "INTEGER",
        "INT": "INTEGER",
        "INT4": "INTEGER",
        "BIGINT": "BIGINT",
        "INT8": "BIGINT",
        "SMALLINT": "SMALLINT",
        "INT2": "SMALLINT",
        "REAL": "REAL",
        "FLOAT4": "REAL",
        "DOUBLE PRECISION": "DOUBLE PRECISION",
        "FLOAT8": "DOUBLE PRECISION",
        "TEXT": "TEXT",
        "DATE": "DATE",
        "TIMESTAMP WITHOUT TIME ZONE": "TIMESTAMP WITHOUT TIME ZONE",
        "TIMESTAMP": "TIMESTAMP WITHOUT TIME ZONE",
    }
    if u in fixed:
        return fixed[u]

    m = re.fullmatch(r"NUMERIC\s*\(\s*(\d{1,4})\s*,\s*(\d{1,4})\s*\)", s, re.I)
    if m:
        p, sc = int(m.group(1)), int(m.group(2))
        if p < 1 or p > 1000 or sc < 0 or sc > p:
            raise ValueError(f"Invalid NUMERIC precision/scale: {raw!r}")
        return f"NUMERIC({p},{sc})"

    m = re.fullmatch(r"NUMERIC\s*\(\s*(\d{1,4})\s*\)", s, re.I)
    if m:
        p = int(m.group(1))
        if p < 1 or p > 1000:
            raise ValueError(f"Invalid NUMERIC precision: {raw!r}")
        return f"NUMERIC({p})"

    m = re.fullmatch(r"VARCHAR\s*\(\s*(\d{1,7})\s*\)", s, re.I)
    if m:
        n = int(m.group(1))
        if n < 1 or n > 10_485_760:
            raise ValueError(f"Invalid VARCHAR length: {raw!r}")
        return f"VARCHAR({n})"

    m = re.fullmatch(r"CHAR\s*\(\s*(\d{1,7})\s*\)", s, re.I)
    if m:
        n = int(m.group(1))
        if n < 1 or n > 10_485_760:
            raise ValueError(f"Invalid CHAR length: {raw!r}")
        return f"CHAR({n})"

    raise ValueError(
        f"Unsupported or invalid PostgreSQL type: {raw!r}. "
        "Use BOOLEAN, INTEGER, BIGINT, REAL, DOUBLE PRECISION, NUMERIC(p,s), "
        "DATE, TIMESTAMP WITHOUT TIME ZONE, TEXT, VARCHAR(n), or CHAR(n)."
    )


def _parse_int_cell(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and not math.isnan(value):
        if value == int(value):
            return int(value)
        return None
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError, OverflowError):
        return None


def combine_base_type_and_size_cells(
    base_raw: Any,
    size_raw: Any,
    scale_raw: Any = None,
) -> str:
    """
    Build a single PostgreSQL type string from a base type cell and optional size/scale cells
    (e.g. type ``VARCHAR`` + size ``255`` → ``VARCHAR(255)``).
    If the base cell already contains ``(``, it is treated as a full type and returned trimmed.
    """
    t = _cell_str_from_raw(base_raw).strip()
    if not t:
        raise ValueError("Empty type cell.")
    if "(" in t:
        return " ".join(t.split())

    u = " ".join(t.split()).upper()
    if u == "CHARACTER VARYING":
        u = "VARCHAR"

    size_s = _cell_str_from_raw(size_raw).strip() if size_raw is not None else ""
    scale_s = _cell_str_from_raw(scale_raw).strip() if scale_raw is not None else ""

    if u in ("VARCHAR", "CHAR"):
        n = _parse_int_cell(size_raw)
        if n is None:
            n = 255
        if n < 1 or n > 10_485_760:
            raise ValueError(f"Invalid length for {u}: {size_raw!r}")
        return f"{u}({n})"

    if u == "NUMERIC" or u.startswith("NUMERIC"):
        if "," in size_s:
            return f"NUMERIC({size_s.replace(' ', '')})"
        p = _parse_int_cell(size_raw)
        sc = _parse_int_cell(scale_raw)
        if p is not None and sc is not None:
            return f"NUMERIC({p},{sc})"
        if p is not None:
            return f"NUMERIC({p})"
        raise ValueError(
            "NUMERIC needs precision (and optional scale) in the size column, "
            "e.g. 18 or 18,6, or use a combined type like NUMERIC(18,6)."
        )

    if size_s or scale_s:
        # Ignore stray size for types that do not use it (INTEGER, TEXT, …)
        pass
    return t


def read_schema_definition_from_sheet(path: Path, sheet_name: str) -> list[tuple[str, str]]:
    """
    Read rows from a sheet whose header row names column name and type columns.

    Standard layout: ``column_name``, ``data_type`` (base type such as ``VARCHAR`` or ``INTEGER``),
    and ``data_size`` (length or precision). Alternatives: one ``data_type`` column only, with full
    types like ``VARCHAR(100)`` in each cell; or legacy headers ``size`` / ``length`` / ``max_length``
    instead of ``data_size``; optional ``scale`` for ``NUMERIC``.
    """
    path = path.resolve()
    headers, rows = read_sheet_rows_raw(path, sheet_name)
    if not headers:
        raise ValueError("Sheet has no header row.")

    keys = [_header_key_label(h) for h in headers]
    name_idx: int | None = None
    for cand in ("column_name", "name", "column", "field_name", "db_column", "field"):
        for i, k in enumerate(keys):
            if k == cand:
                name_idx = i
                break
        if name_idx is not None:
            break

    used: set[int] = set()
    if name_idx is not None:
        used.add(name_idx)

    size_idx: int | None = None
    for cand in ("data_size", "size", "length", "max_length", "char_length", "width"):
        for i, k in enumerate(keys):
            if k == cand and i not in used:
                size_idx = i
                break
        if size_idx is not None:
            break
    if size_idx is not None:
        used.add(size_idx)

    scale_idx: int | None = None
    for cand in ("scale", "decimal_places"):
        for i, k in enumerate(keys):
            if k == cand and i not in used:
                scale_idx = i
                break
        if scale_idx is not None:
            break
    if scale_idx is not None:
        used.add(scale_idx)

    type_idx: int | None = None
    for cand in ("data_type", "type", "pg_type", "postgresql_type", "sql_type", "datatype"):
        for i, k in enumerate(keys):
            if k == cand and i not in used:
                type_idx = i
                break
        if type_idx is not None:
            break

    if name_idx is None or type_idx is None:
        raise ValueError(
            "The first row must include headers for column name and data type, "
            "for example column_name and data_type; for split types, add data_size (or size / length)."
        )
    if name_idx == type_idx:
        raise ValueError("Column name and data type must be two different columns.")

    split_mode = size_idx is not None or scale_idx is not None

    logical_names: list[str] = []
    type_strings: list[str] = []
    size_vals: list[Any] = []
    scale_vals: list[Any] = []
    for row in rows:
        if name_idx >= len(row):
            continue
        if type_idx >= len(row):
            continue
        raw_name = row[name_idx]
        raw_type = row[type_idx]
        name_s = _cell_str_from_raw(raw_name).strip()
        type_s = _cell_str_from_raw(raw_type).strip()
        if not name_s:
            continue
        logical_names.append(name_s)
        type_strings.append(type_s)
        if split_mode:
            sz = row[size_idx] if size_idx is not None and size_idx < len(row) else None
            sc = row[scale_idx] if scale_idx is not None and scale_idx < len(row) else None
            size_vals.append(sz)
            scale_vals.append(sc)
        else:
            size_vals.append(None)
            scale_vals.append(None)

    if not logical_names:
        raise ValueError(
            "No data rows with a non-empty column name. Add rows under the header."
        )

    db_names = unique_column_names(logical_names)
    out: list[tuple[str, str]] = []
    for dbn, ts, sz, sc in zip(
        db_names, type_strings, size_vals, scale_vals, strict=True
    ):
        if not ts and not split_mode:
            raise ValueError(f"Missing data_type for column {dbn!r}.")
        if split_mode:
            combined = combine_base_type_and_size_cells(ts or "", sz, sc)
            out.append((dbn, normalize_pg_type_for_ddl(combined)))
        else:
            if not ts:
                raise ValueError(f"Missing data_type for column {dbn!r}.")
            out.append((dbn, normalize_pg_type_for_ddl(ts)))
    return out


def create_public_table_from_schema_sheet(
    file_path: Path,
    sheet_name: str,
    *,
    host: str,
    port: int,
    user: str,
    password: str,
    dbname: str,
    table_name: str,
) -> dict[str, Any]:
    """
    CREATE TABLE in ``public`` from an Excel sheet defining column names and types.
    Fails if the table already exists.
    """
    if not _validate_table_name(table_name):
        raise ValueError(
            "Invalid table name. Use letters, digits, underscore; start with letter or underscore."
        )
    file_path = file_path.resolve()
    if not file_path.is_file():
        raise ValueError(f"File not found: {file_path}")

    columns = read_schema_definition_from_sheet(file_path, sheet_name)
    parts: list[sql.Composed] = []
    for cname, pg_typ in columns:
        parts.append(sql.SQL("{} {}").format(sql.Identifier(cname), sql.SQL(pg_typ)))
    create_stmt = sql.SQL("CREATE TABLE {} ({})").format(
        sql.Identifier(table_name),
        sql.SQL(", ").join(parts),
    )

    with connect_with_params(
        host=host,
        port=port,
        user=user,
        password=password,
        dbname=dbname,
    ) as conn:
        conn.autocommit = True
        with conn.cursor() as cur:
            try:
                cur.execute(create_stmt)
            except PsycopgError as e:
                raise ValueError(f"CREATE TABLE failed: {e}") from e

    return {
        "table_name": table_name,
        "database": dbname,
        "columns": [[c, t] for c, t in columns],
    }

