"""Rename SOA files: read each file's content to derive a standard name,
move (not copy) renamed files into a parallel '<folder> renamed' folder.
Files that can't be resolved are left in place in the original folder.

Name format:  FY<yy> <Bank/Party> <AccountType> <last 6 digits>
              [ <DD MM YYYY> to <DD MM YYYY> ]   (omitted for a full-FY statement)

Extraction order per file: read file content first; if content doesn't give
a required part, fall back to reading the same clue from the filename. If a
required part still can't be found, the file is left un-renamed in the
original folder and reported.

Only files directly inside the selected folder are processed (no subfolders).
"""

from __future__ import annotations

import csv
import re
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ACCOUNT_TYPES = ("SB", "CA", "CC", "CD", "OD")

_ACCOUNT_TYPE_RE = re.compile(
    r"\b(" + "|".join(ACCOUNT_TYPES) + r")\b", re.IGNORECASE
)
_ACCOUNT_NO_LABEL_RE = re.compile(
    r"a/?c\.?\s*no\.?|account\s*no\.?|account\s*number", re.IGNORECASE
)
_CARD_NO_LABEL_RE = re.compile(r"card\s*no\.?\s*:?", re.IGNORECASE)
_DIGITS_RE = re.compile(r"\d{6,}")

# Common structured header line: "Account <Bank/Party name> <TYPE> <acct no>"
_ACCOUNT_SUMMARY_LINE_RE = re.compile(
    r"account\s+(.+?)\s+(" + "|".join(ACCOUNT_TYPES) + r")\s+(\d{6,})",
    re.IGNORECASE,
)
# Same structured meaning but without the leading "Account" word, and the
# account number split into space-separated digit groups, e.g. "Kotak SB 6130 27418"
_ACCOUNT_LINE_NO_PREFIX_RE = re.compile(
    r"\b([A-Za-z][A-Za-z&]{1,20})\s+(" + "|".join(ACCOUNT_TYPES) + r")\s+(\d{3,6}(?:\s\d{3,6}){0,2})\b",
    re.IGNORECASE,
)
# "HDFC BANK", "For ICICI Bank", "STATE BANK OF INDIA" style mentions
_BANK_NAME_RE = re.compile(r"\b([A-Za-z][A-Za-z&]{1,20})\s+BANK\b", re.IGNORECASE)

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_DATE_NUMERIC_RE = re.compile(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})\b")
_DATE_MONNAME_RE = re.compile(r"\b(\d{1,2})[\s-]([A-Za-z]{3,9})[\s-](\d{2}|\d{4})\b")
# Labeled compact-date statement period, e.g.
# "Transaction Date From: 20230401 ... Transaction Date To: 20231231".
# Deliberately label-anchored (not a blanket \d{8} scan) since statements
# often contain many unrelated 8+ digit reference/transaction numbers.
_PERIOD_FROM_LABEL_RE = re.compile(
    r"(?:transaction\s*date\s*from|statement\s*period\s*from|period\s*from|date\s*from)\s*:?\s*(\d{4})(\d{2})(\d{2})\b",
    re.IGNORECASE,
)
_PERIOD_TO_LABEL_RE = re.compile(
    r"(?:transaction\s*date\s*to|statement\s*period\s*to|period\s*to|date\s*to)\s*:?\s*(\d{4})(\d{2})(\d{2})\b",
    re.IGNORECASE,
)

_MIN_PLAUSIBLE_YEAR = 2000
_MAX_PLAUSIBLE_YEAR = 2100


@dataclass
class RenameSoaReport:
    root: Path
    target: Path
    scanned_count: int = 0
    renamed_count: int = 0
    unresolved: list[str] = field(default_factory=list)
    skipped_paths: list[str] = field(default_factory=list)

    @property
    def skipped_count(self) -> int:
        return len(self.skipped_paths) + len(self.unresolved)


# ---------------------------------------------------------------------------
# File content readers
# ---------------------------------------------------------------------------

def _read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts)


def _cell_to_text(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return str(value)


def _read_excel_text(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [_cell_to_text(c) for c in row if c is not None]
            if cells:
                lines.append(" ".join(cells))
    return "\n".join(lines)


def _read_legacy_xls_text(path: Path) -> str:
    import xlrd

    book = xlrd.open_workbook(str(path))
    lines: list[str] = []
    for sheet in book.sheets():
        for row_idx in range(sheet.nrows):
            cells = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                        cells.append(dt.strftime("%d-%m-%Y"))
                    except Exception:
                        cells.append(str(cell.value))
                else:
                    cells.append(str(cell.value))
            if cells:
                lines.append(" ".join(cells))
    return "\n".join(lines)


def _read_csv_text(path: Path) -> str:
    lines: list[str] = []
    with open(path, newline="", encoding="utf-8", errors="ignore") as f:
        for row in csv.reader(f):
            if row:
                lines.append(" ".join(row))
    return "\n".join(lines)


def _read_image_text(path: Path) -> str:
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return ""
    try:
        return pytesseract.image_to_string(Image.open(path))
    except Exception:
        return ""


def read_file_text(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return _read_pdf_text(path)
        if suffix == ".xls":
            return _read_legacy_xls_text(path)
        if suffix in (".xlsx", ".xlsm"):
            return _read_excel_text(path)
        if suffix == ".csv":
            return _read_csv_text(path)
        if suffix in (".png", ".jpg", ".jpeg", ".bmp", ".tiff"):
            return _read_image_text(path)
    except Exception:
        return ""
    return ""


# ---------------------------------------------------------------------------
# Field extraction from text
# ---------------------------------------------------------------------------

_CREDIT_CARD_HINT_RE = re.compile(r"card\s*no\b|credit\s*card\b", re.IGNORECASE)


def _account_summary_match(text: str) -> re.Match | None:
    """Matches a structured 'Account <Name> <TYPE> <number>' header line, if present."""
    return _ACCOUNT_SUMMARY_LINE_RE.search(text)


def _account_line_no_prefix_match(text: str) -> re.Match | None:
    """Matches '<Bank/Party> <TYPE> <acct no>' with no leading 'Account' word."""
    return _ACCOUNT_LINE_NO_PREFIX_RE.search(text)


def extract_account_type(text: str) -> str | None:
    m = _account_summary_match(text)
    if m:
        return m.group(2).upper()
    m = _account_line_no_prefix_match(text)
    if m:
        return m.group(2).upper()
    m = _ACCOUNT_TYPE_RE.search(text)
    if m:
        return m.group(1).upper()
    if _CREDIT_CARD_HINT_RE.search(text):
        return "CC"
    return None


def extract_last6(text: str) -> str | None:
    m = _account_summary_match(text)
    if m:
        return m.group(3)[-6:]
    m = _account_line_no_prefix_match(text)
    if m:
        digits = re.sub(r"\s+", "", m.group(3))
        if len(digits) >= 4:
            return digits[-6:]
    # Card number is usually shown masked (e.g. "4375 46XX XXXX 3754") — keep
    # the trailing 6 characters (digits and mask X's) as shown, not a plain digit count.
    card_m = _CARD_NO_LABEL_RE.search(text)
    if card_m:
        tail = text[card_m.end(): card_m.end() + 40]
        line_end = tail.find("\n")
        if line_end != -1:
            tail = tail[:line_end]
        cleaned = re.sub(r"[^0-9Xx]", "", tail)
        if len(cleaned) >= 4:
            return cleaned[-6:].upper()
    for m in _ACCOUNT_NO_LABEL_RE.finditer(text):
        tail = text[m.end(): m.end() + 60]
        num_m = _DIGITS_RE.search(tail)
        if num_m:
            digits = num_m.group(0)
            return digits[-6:]
    # fallback: any standalone run of 6+ digits anywhere in the text
    num_m = _DIGITS_RE.search(text)
    if num_m:
        return num_m.group(0)[-6:]
    return None


def extract_bank_party(text: str) -> str | None:
    """Strict content-based bank/party extraction (structured patterns only).

    Deliberately does not fall back to a loose "first substantial line"
    guess here — that heuristic is weaker than a known bank name found in
    the filename, so it's tried only as an absolute last resort in
    _resolve_fields, after the filename fallback has also failed.
    """
    m = _account_summary_match(text)
    if m:
        name = m.group(1).strip()
        if name:
            return name
    m = _account_line_no_prefix_match(text)
    if m:
        name = m.group(1).strip()
        if name:
            return name
    # Only look for a "<Name> BANK" mention near the top of the document
    # (header/logo area). Searching the whole text risks matching an
    # unrelated bank named in a transaction description further down.
    header_window = text[:800]
    bank_m = _BANK_NAME_RE.search(header_window)
    if bank_m:
        return bank_m.group(1).strip()
    return None


def _bank_party_loose_line_guess(text: str) -> str | None:
    """Weak last-resort guess: first substantial text line."""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        letters = sum(1 for c in line if c.isalpha())
        if letters < 3:
            continue
        if len(line) > 60:
            continue
        return line
    return None


def _full_year(yy: int) -> int:
    return 2000 + yy if yy < 100 else yy


def _parse_date_match(kind: str, groups: tuple[str, ...]) -> date | None:
    if kind == "numeric":
        d, mo, y = groups
        try:
            return date(_full_year(int(y)), int(mo), int(d))
        except ValueError:
            return None
    d, mon_name, y = groups
    mo = _MONTHS.get(mon_name.lower()[:3])
    if not mo:
        return None
    try:
        return date(_full_year(int(y)), mo, int(d))
    except ValueError:
        return None


def extract_all_dates(text: str) -> list[date]:
    dates: list[date] = []
    for m in _DATE_NUMERIC_RE.finditer(text):
        d = _parse_date_match("numeric", m.groups())
        if d and _MIN_PLAUSIBLE_YEAR <= d.year <= _MAX_PLAUSIBLE_YEAR:
            dates.append(d)
    for m in _DATE_MONNAME_RE.finditer(text):
        d = _parse_date_match("monname", m.groups())
        if d and _MIN_PLAUSIBLE_YEAR <= d.year <= _MAX_PLAUSIBLE_YEAR:
            dates.append(d)
    return dates


def _labeled_period(text: str) -> tuple[date, date] | None:
    from_m = _PERIOD_FROM_LABEL_RE.search(text)
    to_m = _PERIOD_TO_LABEL_RE.search(text)
    if not from_m or not to_m:
        return None
    try:
        start = date(int(from_m.group(1)), int(from_m.group(2)), int(from_m.group(3)))
        end = date(int(to_m.group(1)), int(to_m.group(2)), int(to_m.group(3)))
    except ValueError:
        return None
    if not (_MIN_PLAUSIBLE_YEAR <= start.year <= _MAX_PLAUSIBLE_YEAR):
        return None
    if not (_MIN_PLAUSIBLE_YEAR <= end.year <= _MAX_PLAUSIBLE_YEAR):
        return None
    return min(start, end), max(start, end)


def extract_period(text: str) -> tuple[date, date] | None:
    labeled = _labeled_period(text)
    if labeled:
        return labeled
    dates = extract_all_dates(text)
    if not dates:
        return None
    return min(dates), max(dates)


def fy_label_for_date(d: date) -> tuple[str, date, date] | None:
    """Returns (fy_no, fy_start, fy_end) for the fiscal year containing ``d``."""
    from gl.fiscal_years.models import FiscalYear

    fy = (
        FiscalYear.objects.filter(start_date__lte=d, end_date__gte=d)
        .order_by("-fy_no")
        .first()
    )
    if fy is None:
        return None
    return fy.fy_no, fy.start_date, fy.end_date


# ---------------------------------------------------------------------------
# Filename fallback extraction (best-effort, same field semantics)
# ---------------------------------------------------------------------------

_KNOWN_BANK_KEYWORDS = [
    "HDFC", "ICICI", "KOTAK", "SBI", "AXIS", "YES BANK", "YES",
    "IDBI", "PNB", "CANARA", "UNION BANK", "INDUSIND", "IDFC",
    "FEDERAL BANK", "RBL", "BANK OF BARODA", "BOB", "DBS", "CITI",
    "STANDARD CHARTERED", "INDIAN BANK", "KARUR VYSYA", "KVB",
    "SOUTH INDIAN BANK", "KARNATAKA BANK", "CITY UNION BANK",
    "AU SMALL FINANCE", "BANDHAN",
]


def _bank_party_from_filename(stem: str) -> str | None:
    upper = stem.upper()
    for keyword in _KNOWN_BANK_KEYWORDS:
        if keyword in upper:
            return keyword
    return None


def _fields_from_filename(stem: str) -> dict:
    out: dict = {}
    # Underscores and hyphens count as "word characters" to regex \b, so
    # "..._CA_..." never matches \bCA\b. Normalize separators to spaces
    # for the word-boundary-based matches below.
    normalized = re.sub(r"[_\-]+", " ", stem)

    m = _ACCOUNT_TYPE_RE.search(normalized)
    if m:
        out["account_type"] = m.group(1).upper()
    num_m = _DIGITS_RE.search(stem)
    if num_m:
        out["last6"] = num_m.group(0)[-6:]
    dates = extract_all_dates(normalized)
    if len(dates) >= 2:
        out["period"] = (min(dates), max(dates))
    fy_m = re.search(r"\bFY\s*-?\s*(\d{2})\b", normalized, re.IGNORECASE)
    if fy_m:
        out["fy_no"] = f"FY{fy_m.group(1)}"
    bank_party = _bank_party_from_filename(stem)
    if bank_party:
        out["bank_party"] = bank_party
    return out


# ---------------------------------------------------------------------------
# Core rename logic
# ---------------------------------------------------------------------------

def _resolve_fields(text: str, stem: str) -> tuple[dict, list[str]]:
    """Resolve fy_no, bank_party, account_type, last6, period (or None).

    Tries file content first, falls back to filename per missing field.
    Returns (fields, missing) where missing lists the field names that
    couldn't be resolved from either source.
    """
    fields: dict = {}
    missing: list[str] = []

    account_type = extract_account_type(text)
    last6 = extract_last6(text)
    bank_party = extract_bank_party(text)
    period = extract_period(text)

    fallback = _fields_from_filename(stem)

    if not account_type:
        account_type = fallback.get("account_type")
    if not last6:
        last6 = fallback.get("last6")
    if not period:
        period = fallback.get("period")
    if not bank_party:
        bank_party = fallback.get("bank_party")
    if not bank_party:
        # Absolute last resort: a weak "first substantial line" guess from
        # content, tried only after both content-strict and filename
        # fallbacks have failed to find a bank/party name.
        bank_party = _bank_party_loose_line_guess(text)

    fy_info = None
    if period:
        fy_info = fy_label_for_date(period[1])
    if not fy_info and "fy_no" in fallback:
        # Filename gave an FY code directly; no start/end to compare against,
        # so treat as non-full-FY (date range, if any, still applies).
        fields["fy_no"] = fallback["fy_no"]
        fields["is_full_fy"] = False
    elif fy_info:
        fy_no, fy_start, fy_end = fy_info
        fields["fy_no"] = fy_no
        fields["is_full_fy"] = period == (fy_start, fy_end)
    else:
        missing.append("fiscal year")

    if not bank_party:
        missing.append("bank/party name")
    else:
        fields["bank_party"] = bank_party

    if not account_type:
        missing.append("account type")
    else:
        fields["account_type"] = account_type

    if not last6:
        missing.append("account number")
    else:
        fields["last6"] = last6

    fields["period"] = period
    return fields, missing


_INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*]')


def _sanitize_filename_part(value: str) -> str:
    cleaned = _INVALID_FILENAME_CHARS_RE.sub("", value)
    cleaned = " ".join(cleaned.split())
    return cleaned.strip(" .")


def _build_new_stem(fields: dict) -> str:
    parts = [fields["fy_no"], fields["bank_party"], fields["account_type"], fields["last6"]]
    if not fields.get("is_full_fy", False) and fields.get("period"):
        start, end = fields["period"]
        parts.append(start.strftime("%d %m %Y"))
        parts.append("to")
        parts.append(end.strftime("%d %m %Y"))
    return _sanitize_filename_part(" ".join(parts))


def _next_available_name(folder: Path, stem: str, suffix: str) -> str:
    base = f"{stem}{suffix}"
    if not (folder / base).exists():
        return base
    n = 2
    while True:
        candidate = f"{stem} v{n}{suffix}"
        if not (folder / candidate).exists():
            return candidate
        n += 1


def rename_soa_files(root: Path, progress_callback=None) -> RenameSoaReport:
    target = root.parent / f"{root.name} renamed"
    target.mkdir(parents=True, exist_ok=True)

    report = RenameSoaReport(root=root, target=target)

    files = [item for item in root.iterdir() if item.is_file()]
    total = len(files)

    for item in files:
        report.scanned_count += 1
        if progress_callback:
            progress_callback(
                "processing", report.scanned_count, total, f"Reading {item.name}..."
            )

        original_name = item.name
        text = read_file_text(item)
        fields, missing = _resolve_fields(text, item.stem)

        if missing:
            report.unresolved.append(
                f"{original_name}: not possible to rename ({', '.join(missing)} not found), left in place"
            )
            continue

        new_stem = _build_new_stem(fields)
        final_name = _next_available_name(target, new_stem, item.suffix)
        destination = target / final_name

        try:
            shutil.move(str(item), str(destination))
        except OSError as exc:
            report.skipped_paths.append(f"{original_name}: {exc}")
            continue
        report.renamed_count += 1

    return report
