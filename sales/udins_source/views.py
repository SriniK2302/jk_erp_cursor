import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook

from sales.udins.models import Udin
from sales.udins.udin_no import normalize_udin

from .forms import UdinSourceHeaderMapForm, UdinSourceImportForm, save_header_map
from .models import UdinSource, UdinSourceHeaderMap
from .workflow import mark_source_row_copied_to_udins


def _norm_header(text):
    s = str(text or "").strip().lower()
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _header_alias_index(header_cells, mapping):
    normalized_headers = [_norm_header(cell) for cell in header_cells]
    out = {}
    for canonical, aliases in mapping.items():
        targets = {_norm_header(canonical)}
        targets.update(_norm_header(alias) for alias in aliases)
        found_idx = None
        for idx, h in enumerate(normalized_headers):
            if h in targets:
                found_idx = idx
                break
        out[canonical] = found_idx
    return out


def _first_header_row(worksheet):
    for row_idx in range(1, min(worksheet.max_row, 40) + 1):
        values = [worksheet.cell(row=row_idx, column=c).value for c in range(1, worksheet.max_column + 1)]
        normalized = {_norm_header(v) for v in values if v is not None}
        if "udin" in normalized and "document_type" in normalized:
            return row_idx, values
    return None, []


def _cell_as_str(cell):
    if cell is None:
        return ""
    return str(cell).strip()


def _to_int_or_none(value):
    s = _cell_as_str(value)
    if s.isdigit():
        return int(s)
    return None


def _source_fields_from_row(*, g, ws, row_number: int, user) -> dict:
    return {
        "s_no": _to_int_or_none(g("s_no")),
        "mrn": g("mrn"),
        "firm": g("firm"),
        "document_type": g("document_type"),
        "document_sub_type": g("document_sub_type"),
        "other_doc": g("other_doc"),
        "document_description": g("document_description"),
        "date_of_signing_of_document": g("date_of_signing_of_document"),
        "ay_fy": g("ay_fy"),
        "created_date_time": g("created_date_time"),
        "remarks": g("remarks"),
        "status": g("status"),
        "particulars_1": g("particulars_1"),
        "figures_values_1": g("figures_values_1"),
        "particulars_2": g("particulars_2"),
        "figures_values_2": g("figures_values_2"),
        "particulars_3": g("particulars_3"),
        "figures_values_3": g("figures_values_3"),
        "particulars_4": g("particulars_4"),
        "figures_values_4": g("figures_values_4"),
        "source_payload": {"sheet": ws.title, "row_number": row_number},
        "imported_by": user,
    }


@login_required
def udins_source(request):
    header_map_row = UdinSourceHeaderMap.get_solo()
    import_form = UdinSourceImportForm()
    map_form = UdinSourceHeaderMapForm(
        initial={"mapping_json": UdinSourceHeaderMapForm.initial_json(header_map_row.mapping_json)}
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "delete":
            row = get_object_or_404(UdinSource, pk=request.POST.get("pk"))
            row.delete()
            messages.success(request, "UDIN source row deleted.")
            return redirect("udins_source")

        if action == "save_mapping":
            map_form = UdinSourceHeaderMapForm(request.POST)
            if map_form.is_valid():
                save_header_map(header_map_row, map_form.cleaned_data["mapping_json"], user=request.user)
                messages.success(request, "Header mapping saved.")
                return redirect("udins_source")
            messages.error(request, "Fix mapping JSON errors and retry.")

        if action == "import_excel":
            import_form = UdinSourceImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                up_file = import_form.cleaned_data["source_file"]
                try:
                    wb = load_workbook(up_file, data_only=True)
                except Exception as exc:
                    messages.error(request, f"Could not read Excel file: {exc}")
                    return redirect("udins_source")

                ws = wb[wb.sheetnames[0]]
                header_row_idx, header_values = _first_header_row(ws)
                if not header_row_idx:
                    messages.error(request, "Could not find UDIN header row in the uploaded file.")
                    return redirect("udins_source")

                col_index = _header_alias_index(header_values, header_map_row.mapping_json)
                created_count = 0
                updated_count = 0
                on_register_count = 0
                duplicate_in_file = 0
                seen_in_file: set[str] = set()
                with transaction.atomic():
                    for r in range(header_row_idx + 1, ws.max_row + 1):
                        row_cells = [
                            ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)
                        ]
                        if not any(_cell_as_str(v) for v in row_cells):
                            continue

                        def g(key):
                            idx = col_index.get(key)
                            if idx is None:
                                return ""
                            return _cell_as_str(row_cells[idx])

                        udin = normalize_udin(g("udin"))
                        if not udin:
                            continue
                        if udin in seen_in_file:
                            duplicate_in_file += 1
                            continue
                        seen_in_file.add(udin)
                        on_register = Udin.objects.filter(udin=udin).exists()
                        fields = _source_fields_from_row(g=g, ws=ws, row_number=r, user=request.user)
                        existing = UdinSource.objects.filter(udin=udin).first()
                        if on_register:
                            on_register_count += 1
                            if existing:
                                for key, value in fields.items():
                                    if key != "imported_by":
                                        setattr(existing, key, value)
                                existing.copied_to_udins = True
                                if existing.copied_on is None:
                                    existing.copied_on = timezone.now()
                                existing.imported_by = request.user
                                existing.save()
                                updated_count += 1
                            else:
                                UdinSource.objects.create(
                                    udin=udin,
                                    copied_to_udins=True,
                                    copied_on=timezone.now(),
                                    **fields,
                                )
                                created_count += 1
                            continue
                        if existing:
                            for key, value in fields.items():
                                if key != "imported_by":
                                    setattr(existing, key, value)
                            existing.copied_to_udins = False
                            existing.copied_on = None
                            existing.imported_by = request.user
                            existing.save()
                            updated_count += 1
                        else:
                            UdinSource.objects.create(udin=udin, **fields)
                            created_count += 1

                parts = [f"Created: {created_count}"]
                if updated_count:
                    parts.append(f"Updated in source: {updated_count}")
                if on_register_count:
                    parts.append(
                        f"Also on UDINs register (kept in source, marked copied): {on_register_count}"
                    )
                if duplicate_in_file:
                    parts.append(f"Skipped (duplicate in file): {duplicate_in_file}")
                messages.success(request, "UDIN source import complete. " + ", ".join(parts) + ".")
                return redirect("udins_source")
            messages.error(request, "Please choose a valid Excel file.")

    copy_filter = (request.GET.get("copy_status") or "pending").strip().lower()
    if copy_filter not in {"all", "pending", "copied"}:
        copy_filter = "pending"

    rows_qs = UdinSource.objects.select_related("imported_by")
    if copy_filter == "pending":
        rows_qs = rows_qs.filter(copied_to_udins=False)
    elif copy_filter == "copied":
        rows_qs = rows_qs.filter(copied_to_udins=True)

    rows = list(rows_qs.all())
    return render(
        request,
        "udins_source/udins_source.html",
        {
            "rows": rows,
            "import_form": import_form,
            "map_form": map_form,
            "copy_filter": copy_filter,
        },
    )
