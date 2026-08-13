from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.test import Client as HttpClient
from django.test import SimpleTestCase, TestCase
from django.urls import reverse

from config.models import ChartOfAccount, SalesLedgerSettings
from gl.fiscal_years.models import FiscalYear
from gl.journal.models import GlHeader, GlLine, TbTable, TbTableMonth
from sales.client_classifications.models import ClientClassification
from sales.clients.models import Client
from sales.invoices.sales_gl_posting import bulk_post_fresh_invoices_to_gl
from sales.invoices.amount_words import rupees_in_words
from sales.invoices.narration_build import (
    header_narration_from_udin_rows,
    narration_suggestion_for_udin,
    reload_narration_templates,
)
from sales.services.models import Service
from sales.udins.models import Udin

from sales.invoices.gstr1_export import gstr1_export_http_response
from sales.invoices.gstr1_invoice_list import (
    compute_gstr1_invoice_list,
    window_for_fy_month,
)
from sales.invoices.monthly_invoice_summary import compute_monthly_invoice_summary
from sales.invoices.models import Invoice, InvoiceLine, InvoiceStatus
from sales.invoices.sales_ledger_tb import compute_sales_ledger_tb


class RupeesInWordsTests(SimpleTestCase):
    def test_three_thousand_five_forty(self):
        self.assertEqual(
            rupees_in_words(3540),
            "Rupees Three Thousand Five Hundred Forty only",
        )

    def test_zero(self):
        self.assertEqual(rupees_in_words(Decimal("0")), "Rupees Zero only")


class NarrationSuggestionTests(TestCase):
    def setUp(self):
        reload_narration_templates()
        User = get_user_model()
        self.user = User.objects.create_user(
            username="nar_user",
            password="pass12345",
        )
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="NarTest",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="Narration Test Co",
            client_short_name="NTC",
            client_code="NTC1",
            classification=self.classification,
            is_active=True,
            created_by=self.user,
        )
        self.svc_cert = Service.objects.create(
            service_desc="Certification",
            service_code="CERT",
            created_by=self.user,
        )

    def test_certification_template_uses_service_remarks(self):
        udin = Udin.objects.create(
            udin="NAR-UDIN-001",
            client=self.client_row,
            service=self.svc_cert,
            service_remarks="FY24 ESI Form 5",
            inv_tv_amount=Decimal("100.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        text = narration_suggestion_for_udin(udin)
        self.assertIn("certificate", text.lower())
        self.assertIn("FY24 ESI Form 5", text)

    def test_certification_template_matches_longer_service_desc(self):
        """by_service_desc key 'Certification' applies when master service text is extended."""
        reload_narration_templates()
        svc = Service.objects.create(
            service_desc="Certification - post merger solvency",
            service_code="CX01",
            created_by=self.user,
        )
        udin = Udin.objects.create(
            udin="NAR-UDIN-004",
            client=self.client_row,
            service=svc,
            service_remarks="post merger solvency",
            inv_tv_amount=Decimal("1.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        text = narration_suggestion_for_udin(udin)
        self.assertIn("certificate", text.lower())
        self.assertIn("post merger solvency", text)

    def test_statutory_audit_template_case_insensitive_desc(self):
        reload_narration_templates()
        svc = Service.objects.create(
            service_desc="statutory audit",
            service_code="SAUD",
            created_by=self.user,
        )
        udin = Udin.objects.create(
            udin="NAR-UDIN-002",
            client=self.client_row,
            service=svc,
            service_remarks="",
            ay_fy="FY26",
            inv_tv_amount=Decimal("1.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        text = narration_suggestion_for_udin(udin)
        self.assertIn("Statutory Audit", text)
        self.assertIn("service", text.lower())
        self.assertIn("conducting", text.lower())
        self.assertIn("issuing", text.lower())

    def test_tax_audit_template_excludes_statutory_wording(self):
        reload_narration_templates()
        svc = Service.objects.create(
            service_desc="tax audit",
            service_code="TAUD",
            created_by=self.user,
        )
        udin = Udin.objects.create(
            udin="NAR-UDIN-003",
            client=self.client_row,
            service=svc,
            ay_fy="FY27",
            inv_tv_amount=Decimal("1.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        text = narration_suggestion_for_udin(udin)
        self.assertIn("Tax Audit", text)
        self.assertNotIn("Statutory Audit", text)
        self.assertIn("service", text.lower())

    def test_header_narration_from_two_udins_joins_distinct_suggestions(self):
        reload_narration_templates()
        svc_cert = Service.objects.create(
            service_desc="Certification",
            service_code="C2",
            created_by=self.user,
        )
        svc_tax = Service.objects.create(
            service_desc="Tax Audit",
            service_code="T2",
            created_by=self.user,
        )
        u1 = Udin.objects.create(
            udin="NAR-UDIN-H1",
            client=self.client_row,
            service=svc_cert,
            service_remarks="Scope A",
            inv_tv_amount=Decimal("100.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        u2 = Udin.objects.create(
            udin="NAR-UDIN-H2",
            client=self.client_row,
            service=svc_tax,
            service_remarks="",
            inv_tv_amount=Decimal("50.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        rows = [(u1, "line a", Decimal("100")), (u2, "line b", Decimal("50"))]
        text = header_narration_from_udin_rows(rows)
        self.assertIn("certificate", text.lower())
        self.assertIn("Tax Audit", text)


class CreateInvoiceFromUdinTests(TestCase):
    def setUp(self):
        reload_narration_templates()
        User = get_user_model()
        self.user = User.objects.create_user(username="bulk_inv_user", password="pass12345")
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="BulkInv",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="Bulk Inv Co",
            client_short_name="BIC",
            client_code="BIC1",
            classification=self.classification,
            is_active=True,
            created_by=self.user,
        )
        self.svc = Service.objects.create(
            service_desc="Certification",
            service_code="BCER",
            created_by=self.user,
        )
        FiscalYear.objects.create(
            fy_no="FY99",
            start_date=date(2098, 4, 1),
            end_date=date(2099, 3, 31),
            created_by=self.user,
        )

    def test_create_invoice_from_udin_ok(self):
        from sales.invoices.invoice_from_udin import create_invoice_from_udin

        udin = Udin.objects.create(
            udin="BULK-UDIN-001",
            client=self.client_row,
            service=self.svc,
            ay_fy="FY99",
            inv_tv_amount=Decimal("1000.00"),
            inv_date=date(2026, 6, 1),
            is_invoiced=False,
            created_by=self.user,
        )
        inv, err = create_invoice_from_udin(user=self.user, udin=udin)
        self.assertIsNone(err)
        self.assertIsNotNone(inv)
        assert inv is not None
        self.assertEqual(inv.invoice_date, date(2026, 6, 1))
        self.assertEqual(inv.inv_taxable_value, Decimal("1000.00"))
        inv.refresh_from_db()
        self.assertTrue((inv.narration or "").strip(), "auto-created invoice should persist header narration")
        udin.refresh_from_db()
        self.assertTrue(udin.is_invoiced)
        self.assertEqual(udin.inv_no, inv.invoice_no)

    def test_create_invoice_from_udin_rejects_unknown_fy(self):
        from sales.invoices.invoice_from_udin import create_invoice_from_udin

        udin = Udin.objects.create(
            udin="BULK-UDIN-002",
            client=self.client_row,
            service=self.svc,
            ay_fy="FYZZ",
            inv_tv_amount=Decimal("10.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        inv, err = create_invoice_from_udin(user=self.user, udin=udin)
        self.assertIsNone(inv)
        self.assertIsNotNone(err)
        assert err is not None
        self.assertIn("Service FY", err)

    def test_invoice_readiness_issues_lists_missing_fields(self):
        from sales.invoices.invoice_from_udin import invoice_readiness_issues

        issues = invoice_readiness_issues(
            client_id=None,
            service_id=self.svc.pk,
            ay_fy="2026-2027",
            inv_tv_amount=None,
        )
        self.assertIn("Set Client.", issues)
        self.assertIn("Set Inv TV amt.", issues)
        self.assertTrue(any("Service FY" in item for item in issues))

    def test_invoice_readiness_issues_ready(self):
        from sales.invoices.invoice_from_udin import invoice_readiness_issues

        issues = invoice_readiness_issues(
            client_id=self.client_row.pk,
            service_id=self.svc.pk,
            ay_fy="FY99",
            inv_tv_amount=Decimal("500.00"),
        )
        self.assertEqual(issues, [])


class InvoiceFormViewTests(TestCase):
    def setUp(self):
        reload_narration_templates()
        User = get_user_model()
        self.user = User.objects.create_user(
            username="inv_form_user", password="pass12345", is_superuser=True
        )
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="InvForm",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="Inv Form Co",
            client_short_name="IFC",
            client_code="IFC1",
            classification=self.classification,
            is_active=True,
            created_by=self.user,
        )
        self.svc = Service.objects.create(
            service_desc="Certification",
            service_code="IFCR",
            created_by=self.user,
        )
        self.fy = FiscalYear.objects.create(
            fy_no="FY98",
            start_date=date(2097, 4, 1),
            end_date=date(2098, 3, 31),
            created_by=self.user,
        )
        self.udin = Udin.objects.create(
            udin="FORM-UDIN-001",
            client=self.client_row,
            service=self.svc,
            ay_fy="FY98",
            inv_tv_amount=Decimal("500.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        self.client.force_login(self.user)

    def _post_data(self, invoice_no, *, initial_forms="0", udin=None):
        return {
            "invoice_date": "2097-05-01",
            "invoice_no": invoice_no,
            "fiscal_year": str(self.fy.pk),
            "status": "fresh",
            "narration": "Fee for testing.",
            "maps-TOTAL_FORMS": "1",
            "maps-INITIAL_FORMS": initial_forms,
            "maps-MIN_NUM_FORMS": "0",
            "maps-MAX_NUM_FORMS": "1000",
            "maps-0-udin": str((udin or self.udin).pk),
            "maps-0-service_desc": "Certification",
            "maps-0-line_amount": "500.00",
        }

    def test_create_page_udin_meta_includes_inv_date(self):
        from sales.invoices.views import _udin_choice_meta

        self.udin.inv_date = date(2026, 6, 1)
        self.udin.save()
        meta = _udin_choice_meta(Udin.objects.filter(pk=self.udin.pk))
        self.assertEqual(meta[str(self.udin.pk)]["inv_date"], "2026-06-01")

    def test_create_then_resubmit_edit_with_stale_state_succeeds(self):
        response = self.client.post(reverse("invoice_create"), self._post_data("FY98-IFC-001"))
        self.assertEqual(response.status_code, 302)
        inv = Invoice.objects.get(client=self.client_row, invoice_no="FY98-IFC-001")
        edit_url = reverse("invoice_edit", kwargs={"pk": inv.pk})
        # Two consecutive saves: the second mimics a stale-page resubmit after
        # the map rows were deleted and recreated by the first save.
        for _ in range(2):
            response = self.client.post(edit_url, self._post_data("FY98-IFC-001", initial_forms="1"))
            self.assertEqual(response.status_code, 302)
        inv.refresh_from_db()
        self.assertEqual(inv.inv_taxable_value, Decimal("500.00"))
        self.assertEqual(inv.inv_udin_maps.count(), 1)

    def test_duplicate_invoice_no_shows_form_error_not_500(self):
        response = self.client.post(reverse("invoice_create"), self._post_data("FY98-IFC-002"))
        self.assertEqual(response.status_code, 302)
        second_udin = Udin.objects.create(
            udin="FORM-UDIN-002",
            client=self.client_row,
            service=self.svc,
            ay_fy="FY98",
            inv_tv_amount=Decimal("700.00"),
            is_invoiced=False,
            created_by=self.user,
        )
        response = self.client.post(
            reverse("invoice_create"),
            self._post_data("FY98-IFC-002", udin=second_udin),
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already exists")
        self.assertEqual(
            Invoice.objects.filter(client=self.client_row, invoice_no="FY98-IFC-002").count(),
            1,
        )


class SalesLedgerTbTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="sl_tb_user", password="pass12345")
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="SlTb",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="SL TB Co",
            client_short_name="SLTB",
            client_code="SLT1",
            classification=self.classification,
            is_active=True,
            created_by=self.user,
        )
        self.svc = Service.objects.create(
            service_desc="Cert",
            service_code="SLC1",
            created_by=self.user,
        )
        self.fy = FiscalYear.objects.create(
            fy_no="FYTB",
            start_date=date(2098, 4, 1),
            end_date=date(2099, 3, 31),
            created_by=self.user,
        )

    def test_compute_tb_balances_gst_invoice(self):
        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2098, 6, 1),
            invoice_no="FYTB-SLT1-001",
            inv_taxable_value=Decimal("1000.00"),
            taxes=Decimal("180.00"),
            inv_gross=Decimal("1180.00"),
            created_by=self.user,
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=1,
            line_type="Service",
            line_base_amount=Decimal("1000.00"),
            percentage=Decimal("100"),
            item_amount=Decimal("1000.00"),
            line_description="Fee",
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=2,
            line_type="CGST",
            line_base_amount=Decimal("1000.00"),
            percentage=Decimal("9"),
            item_amount=Decimal("90.00"),
            line_description="",
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=3,
            line_type="SGST",
            line_base_amount=Decimal("1000.00"),
            percentage=Decimal("9"),
            item_amount=Decimal("90.00"),
            line_description="",
        )
        tb = compute_sales_ledger_tb(self.fy)
        self.assertTrue(tb["balanced"])
        self.assertEqual(tb["grand_total_dr"], Decimal("1180.00"))
        self.assertEqual(tb["grand_total_cr"], Decimal("1180.00"))
        self.assertEqual(tb["period_invoice_count"], 1)
        self.assertEqual(tb["opening_invoice_count"], 0)
        self.assertEqual(len(tb["period"]["client_rows"]), 1)
        self.assertEqual(tb["period"]["client_rows"][0]["invoice_count"], 1)
        self.assertEqual(tb["period"]["client_rows"][0]["debit"], Decimal("1180.00"))

    def test_tb_uses_invoice_date_not_service_fy_fk(self):
        """Ledger TB window is invoice_date vs FY dates; Service FY on invoice is ignored."""
        fy_other = FiscalYear.objects.create(
            fy_no="FYOT",
            start_date=date(2090, 4, 1),
            end_date=date(2091, 3, 31),
            created_by=self.user,
        )
        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=fy_other,
            invoice_date=date(2098, 7, 1),
            invoice_no="FYTB-OTHERFY-001",
            inv_taxable_value=Decimal("100.00"),
            taxes=Decimal("18.00"),
            inv_gross=Decimal("118.00"),
            created_by=self.user,
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=1,
            line_type="Service",
            line_base_amount=Decimal("100.00"),
            percentage=Decimal("100"),
            item_amount=Decimal("100.00"),
            line_description="",
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=2,
            line_type="CGST",
            line_base_amount=Decimal("100.00"),
            percentage=Decimal("9"),
            item_amount=Decimal("9.00"),
            line_description="",
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=3,
            line_type="SGST",
            line_base_amount=Decimal("100.00"),
            percentage=Decimal("9"),
            item_amount=Decimal("9.00"),
            line_description="",
        )
        tb = compute_sales_ledger_tb(self.fy)
        self.assertEqual(tb["period_invoice_count"], 1)
        self.assertEqual(tb["opening_invoice_count"], 0)
        self.assertTrue(tb["balanced"])

    def test_opening_bf_when_period_has_no_invoices(self):
        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2098, 1, 15),
            invoice_no="FYTB-PRIOR-001",
            inv_taxable_value=Decimal("200.00"),
            taxes=Decimal("36.00"),
            inv_gross=Decimal("236.00"),
            created_by=self.user,
        )
        for line_no, ltype, pct, item in [
            (1, "Service", Decimal("100"), Decimal("200.00")),
            (2, "CGST", Decimal("9"), Decimal("18.00")),
            (3, "SGST", Decimal("9"), Decimal("18.00")),
        ]:
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=line_no,
                line_type=ltype,
                line_base_amount=Decimal("200.00") if line_no == 1 else Decimal("200.00"),
                percentage=pct,
                item_amount=item,
                line_description="",
            )
        tb = compute_sales_ledger_tb(self.fy)
        self.assertEqual(tb["period_invoice_count"], 0)
        self.assertEqual(tb["opening_invoice_count"], 1)
        self.assertTrue(tb["balanced"])
        self.assertEqual(tb["grand_total_dr"], Decimal("236.00"))
        self.assertGreater(tb["opening"]["balance_carried_credit"], 0)

    def _minimal_gst_invoice(self, *, invoice_no, invoice_date, inv_gross, status):
        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=invoice_date,
            invoice_no=invoice_no,
            inv_taxable_value=Decimal("100.00"),
            taxes=Decimal("18.00"),
            inv_gross=inv_gross,
            status=status,
            created_by=self.user,
        )
        for line_no, ltype, pct, item in [
            (1, "Service", Decimal("100"), Decimal("100.00")),
            (2, "CGST", Decimal("9"), Decimal("9.00")),
            (3, "SGST", Decimal("9"), Decimal("9.00")),
        ]:
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=line_no,
                line_type=ltype,
                line_base_amount=Decimal("100.00"),
                percentage=pct,
                item_amount=item,
                line_description="",
            )
        return inv

    def test_tb_status_filter_all_authorised_fresh(self):
        self._minimal_gst_invoice(
            invoice_no="FYTB-ST-001",
            invoice_date=date(2098, 6, 1),
            inv_gross=Decimal("118.00"),
            status=InvoiceStatus.FRESH,
        )
        self._minimal_gst_invoice(
            invoice_no="FYTB-ST-002",
            invoice_date=date(2098, 6, 2),
            inv_gross=Decimal("118.00"),
            status=InvoiceStatus.AUTHORISED,
        )
        tb_all = compute_sales_ledger_tb(self.fy, invoice_status="all")
        self.assertEqual(tb_all["period_invoice_count"], 2)
        self.assertEqual(tb_all["grand_total_dr"], Decimal("236.00"))
        self.assertTrue(tb_all["balanced"])

        tb_auth = compute_sales_ledger_tb(self.fy, invoice_status="authorised")
        self.assertEqual(tb_auth["period_invoice_count"], 1)
        self.assertEqual(tb_auth["grand_total_dr"], Decimal("118.00"))
        self.assertTrue(tb_auth["balanced"])

        tb_fresh = compute_sales_ledger_tb(self.fy, invoice_status="fresh")
        self.assertEqual(tb_fresh["period_invoice_count"], 1)
        self.assertEqual(tb_fresh["grand_total_dr"], Decimal("118.00"))
        self.assertTrue(tb_fresh["balanced"])


class Gstr1InvoiceListTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="gstr1_user", password="pass12345")
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="Gstr1",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="GSTR One Co Ltd",
            client_short_name="G1CO",
            client_code="G1C1",
            classification=self.classification,
            is_active=True,
            billing_gstn="29ABCDE1234F1Z5",
            created_by=self.user,
        )
        self.svc = Service.objects.create(
            service_desc="Cert",
            service_code="G1SV",
            created_by=self.user,
        )
        self.fy = FiscalYear.objects.create(
            fy_no="FYG1",
            start_date=date(2098, 4, 1),
            end_date=date(2099, 3, 31),
            created_by=self.user,
        )

    def test_gstr1_aggregates_multiple_service_lines_per_invoice(self):
        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2098, 6, 15),
            invoice_no="FYG1-MULTI-001",
            inv_taxable_value=Decimal("300.00"),
            taxes=Decimal("54.00"),
            inv_gross=Decimal("354.00"),
            narration="Two fee lines",
            created_by=self.user,
        )
        lines = [
            (1, "Service", Decimal("100"), Decimal("100.00")),
            (2, "Service", Decimal("100"), Decimal("200.00")),
            (3, "CGST", Decimal("9"), Decimal("27.00")),
            (4, "SGST", Decimal("9"), Decimal("27.00")),
        ]
        for line_no, ltype, pct, item in lines:
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=line_no,
                line_type=ltype,
                line_base_amount=Decimal("300.00") if ltype == "Service" else Decimal("300.00"),
                percentage=pct,
                item_amount=item,
                line_description="",
            )
        w = window_for_fy_month(self.fy, month_first=date(2098, 6, 1), ytd=False)
        rows = compute_gstr1_invoice_list(w.date_from, w.date_to, invoice_status="all")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["taxable_value"], Decimal("300.00"))
        self.assertEqual(rows[0]["cgst"], Decimal("27.00"))
        self.assertEqual(rows[0]["sgst"], Decimal("27.00"))
        self.assertEqual(rows[0]["inv_gross"], Decimal("354.00"))
        self.assertEqual(rows[0]["client_gstn"], "29ABCDE1234F1Z5")
        self.assertEqual(rows[0]["client_name"], "GSTR One Co Ltd")

    def test_monthly_summary_aggregates_by_month_and_fy_total(self):
        for month, day, no in [(5, 10, "FYG1-MAY"), (6, 15, "FYG1-JUN")]:
            inv = Invoice.objects.create(
                client=self.client_row,
                service=self.svc,
                fiscal_year=self.fy,
                invoice_date=date(2098, month, day),
                invoice_no=no,
                inv_taxable_value=Decimal("100.00"),
                taxes=Decimal("18.00"),
                inv_gross=Decimal("118.00"),
                created_by=self.user,
            )
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=1,
                line_type="Service",
                line_base_amount=Decimal("100.00"),
                percentage=Decimal("100"),
                item_amount=Decimal("100.00"),
                line_description="",
            )
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=2,
                line_type="CGST",
                line_base_amount=Decimal("100.00"),
                percentage=Decimal("9"),
                item_amount=Decimal("9.00"),
                line_description="",
            )
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=3,
                line_type="SGST",
                line_base_amount=Decimal("100.00"),
                percentage=Decimal("9"),
                item_amount=Decimal("9.00"),
                line_description="",
            )
        month_rows, total = compute_monthly_invoice_summary(self.fy, invoice_status="all")
        june = next(r for r in month_rows if r["month_label"] == "Jun 2098")
        self.assertEqual(june["invoice_count"], 1)
        self.assertEqual(june["taxable_value"], Decimal("100.00"))
        self.assertEqual(total["invoice_count"], 2)
        self.assertEqual(total["inv_gross"], Decimal("236.00"))

    def test_gstr1_ytd_includes_prior_months_in_same_fy(self):
        for day, no, gross in [
            (10, "FYG1-MAY", Decimal("118.00")),
            (20, "FYG1-JUN", Decimal("118.00")),
        ]:
            inv = Invoice.objects.create(
                client=self.client_row,
                service=self.svc,
                fiscal_year=self.fy,
                invoice_date=date(2098, 5 if no.endswith("MAY") else 6, day),
                invoice_no=no,
                inv_taxable_value=Decimal("100.00"),
                taxes=Decimal("18.00"),
                inv_gross=gross,
                created_by=self.user,
            )
            for line_no, ltype, pct, item in [
                (1, "Service", Decimal("100"), Decimal("100.00")),
                (2, "CGST", Decimal("9"), Decimal("9.00")),
                (3, "SGST", Decimal("9"), Decimal("9.00")),
            ]:
                InvoiceLine.objects.create(
                    invoice=inv,
                    line_no=line_no,
                    line_type=ltype,
                    line_base_amount=Decimal("100.00"),
                    percentage=pct,
                    item_amount=item,
                    line_description="",
                )

        w_month = window_for_fy_month(self.fy, month_first=date(2098, 6, 1), ytd=False)
        rows_m = compute_gstr1_invoice_list(w_month.date_from, w_month.date_to, invoice_status="all")
        self.assertEqual(len(rows_m), 1)
        self.assertEqual(rows_m[0]["invoice_no"], "FYG1-JUN")

        w_ytd = window_for_fy_month(self.fy, month_first=date(2098, 6, 1), ytd=True)
        rows_y = compute_gstr1_invoice_list(w_ytd.date_from, w_ytd.date_to, invoice_status="all")
        self.assertEqual(len(rows_y), 2)
        nos = {r["invoice_no"] for r in rows_y}
        self.assertEqual(nos, {"FYG1-MAY", "FYG1-JUN"})

    def test_gstr1_export_csv_has_headers_and_row(self):
        w = window_for_fy_month(self.fy, month_first=date(2098, 6, 1), ytd=False)
        rows = compute_gstr1_invoice_list(w.date_from, w.date_to, invoice_status="all")
        resp = gstr1_export_http_response(
            report_rows=rows,
            export_fmt="csv",
            filename_base="test_export",
        )
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8-sig")
        self.assertIn("Client GSTN", body)
        self.assertIn("Inv No", body)

    def test_gstr1_export_xlsx_amounts_and_date_are_excel_numeric_types(self):
        from openpyxl import load_workbook

        inv = Invoice.objects.create(
            client=self.client_row,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2098, 6, 15),
            invoice_no="FYG1-XLSX",
            inv_taxable_value=Decimal("100.00"),
            taxes=Decimal("18.00"),
            inv_gross=Decimal("118.00"),
            narration="xlsx export",
            created_by=self.user,
        )
        for line_no, ltype, pct, item in [
            (1, "Service", Decimal("100"), Decimal("100.00")),
            (2, "CGST", Decimal("9"), Decimal("9.00")),
            (3, "SGST", Decimal("9"), Decimal("9.00")),
        ]:
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=line_no,
                line_type=ltype,
                line_base_amount=Decimal("100.00"),
                percentage=pct,
                item_amount=item,
                line_description="",
            )
        w = window_for_fy_month(self.fy, month_first=date(2098, 6, 1), ytd=False)
        rows = compute_gstr1_invoice_list(w.date_from, w.date_to, invoice_status="all")
        self.assertEqual(len(rows), 1)
        resp = gstr1_export_http_response(
            report_rows=rows,
            export_fmt="xlsx",
            filename_base="test_xlsx",
        )
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        c_date = ws.cell(row=2, column=4)
        c_taxable = ws.cell(row=2, column=6)
        self.assertEqual(c_date.data_type, "d")
        self.assertEqual(c_taxable.data_type, "n")
        dv = c_date.value
        if isinstance(dv, datetime):
            dv = dv.date()
        self.assertEqual(dv, date(2098, 6, 15))
        self.assertAlmostEqual(float(c_taxable.value), 100.0)


class SalesGlPostingTests(TestCase):
    """Bulk authorise + GL journal from Sales ledger COA settings."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="gl_sales_user", password="pass12345")
        self.user.is_superuser = True
        self.user.save()
        self.classification, _ = ClientClassification.objects.get_or_create(
            classification_name="GlPost",
            defaults={"created_by": self.user},
        )
        self.client_row = Client.objects.create(
            client_name="GL Post Co",
            client_short_name="GLPC",
            client_code="GLP1",
            classification=self.classification,
            is_active=True,
            created_by=self.user,
        )
        self.client_igst = Client.objects.create(
            client_name="GL IGST Co",
            client_short_name="GLIG",
            client_code="GLI1",
            classification=self.classification,
            is_active=True,
            invoice_tax_type=Client.INVOICE_TAX_IGST,
            created_by=self.user,
        )
        self.svc = Service.objects.create(
            service_desc="Audit",
            service_code="GLA1",
            created_by=self.user,
        )
        self.fy = FiscalYear.objects.create(
            fy_no="FYGL",
            start_date=date(2099, 4, 1),
            end_date=date(2100, 3, 31),
            created_by=self.user,
        )
        self.coa_recv = ChartOfAccount.objects.create(
            account_name="Trade Receivables",
            account_code="GL1200",
            plbs=ChartOfAccount.PLBS_BS,
            plbs_type=ChartOfAccount.TYPE_ASSET,
            created_by=self.user,
        )
        self.coa_fee = ChartOfAccount.objects.create(
            account_name="Professional Fees",
            account_code="GL5000",
            plbs=ChartOfAccount.PLBS_PL,
            plbs_type=ChartOfAccount.TYPE_INCOME,
            created_by=self.user,
        )
        self.coa_cgst = ChartOfAccount.objects.create(
            account_name="CGST Output",
            account_code="GL2210",
            plbs=ChartOfAccount.PLBS_BS,
            plbs_type=ChartOfAccount.TYPE_LIABILITY,
            created_by=self.user,
        )
        self.coa_sgst = ChartOfAccount.objects.create(
            account_name="SGST Output",
            account_code="GL2211",
            plbs=ChartOfAccount.PLBS_BS,
            plbs_type=ChartOfAccount.TYPE_LIABILITY,
            created_by=self.user,
        )
        self.coa_igst = ChartOfAccount.objects.create(
            account_name="IGST Output",
            account_code="GL2212",
            plbs=ChartOfAccount.PLBS_BS,
            plbs_type=ChartOfAccount.TYPE_LIABILITY,
            created_by=self.user,
        )
        sls = SalesLedgerSettings.get_solo()
        sls.sales_ledger_control_account = self.coa_recv
        sls.service_income_account = self.coa_fee
        sls.cgst_output_account = self.coa_cgst
        sls.sgst_output_account = self.coa_sgst
        sls.igst_output_account = self.coa_igst
        sls.save()

    def _gst_invoice(self, *, client, invoice_no, gross=Decimal("1180.00")):
        inv = Invoice.objects.create(
            client=client,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2099, 5, 1),
            invoice_no=invoice_no,
            inv_taxable_value=Decimal("1000.00"),
            taxes=Decimal("180.00"),
            inv_gross=gross,
            status=InvoiceStatus.FRESH,
            created_by=self.user,
        )
        for line_no, ltype, pct, item in [
            (1, "Service", Decimal("100"), Decimal("1000.00")),
            (2, "CGST", Decimal("9"), Decimal("90.00")),
            (3, "SGST", Decimal("9"), Decimal("90.00")),
        ]:
            InvoiceLine.objects.create(
                invoice=inv,
                line_no=line_no,
                line_type=ltype,
                line_base_amount=Decimal("1000.00"),
                percentage=pct,
                item_amount=item,
                line_description="",
            )
        return inv

    def test_bulk_post_creates_authorised_gl_and_invoice(self):
        inv = self._gst_invoice(client=self.client_row, invoice_no="FYGL-001")
        n, errs = bulk_post_fresh_invoices_to_gl(invoice_pks=[inv.pk], user=self.user)
        self.assertEqual(n, 1)
        self.assertEqual(errs, [])
        inv.refresh_from_db()
        self.assertEqual(inv.status, InvoiceStatus.AUTHORISED)
        self.assertIsNotNone(inv.posted_gl_header_id)
        hdr = inv.posted_gl_header
        self.assertEqual(hdr.status, GlHeader.Status.AUTHORISED)
        self.assertEqual(hdr.source, GlHeader.Source.SALES)
        self.assertEqual(hdr.tran_date, inv.invoice_date)
        self.assertTrue(hdr.tran_id.startswith("Sales-"))
        self.assertIn(inv.invoice_no, hdr.narration)
        lines = list(hdr.lines.order_by("line_no"))
        self.assertEqual(len(lines), 4)
        self.assertEqual(lines[0].account_id, self.coa_recv.account_code)
        self.assertEqual(lines[0].amount, Decimal("1180.00"))
        self.assertEqual(lines[1].amount, Decimal("-1000.00"))
        self.assertEqual(lines[2].amount, Decimal("-90.00"))
        self.assertEqual(lines[3].amount, Decimal("-90.00"))
        net = sum(GlLine.objects.filter(header=hdr).values_list("amount", flat=True))
        self.assertEqual(net, Decimal("0"))
        self.assertEqual(TbTable.objects.filter(fiscal_year=self.fy).count(), 4)
        recv_tb = TbTable.objects.get(
            fiscal_year=self.fy, account_code=self.coa_recv.account_code
        )
        self.assertEqual(recv_tb.amount, Decimal("1180.00"))
        self.assertEqual(
            TbTableMonth.objects.filter(fiscal_year=self.fy).count(),
            4,
        )

    def test_bulk_post_igst_invoice_three_lines(self):
        inv = Invoice.objects.create(
            client=self.client_igst,
            service=self.svc,
            fiscal_year=self.fy,
            invoice_date=date(2099, 5, 2),
            invoice_no="FYGL-IG-001",
            inv_taxable_value=Decimal("1000.00"),
            taxes=Decimal("180.00"),
            inv_gross=Decimal("1180.00"),
            status=InvoiceStatus.FRESH,
            created_by=self.user,
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=1,
            line_type="Service",
            line_base_amount=Decimal("1000.00"),
            percentage=Decimal("100"),
            item_amount=Decimal("1000.00"),
            line_description="",
        )
        InvoiceLine.objects.create(
            invoice=inv,
            line_no=2,
            line_type="IGST",
            line_base_amount=Decimal("1000.00"),
            percentage=Decimal("18"),
            item_amount=Decimal("180.00"),
            line_description="",
        )
        n, errs = bulk_post_fresh_invoices_to_gl(invoice_pks=[inv.pk], user=self.user)
        self.assertEqual(n, 1)
        self.assertEqual(errs, [])
        inv.refresh_from_db()
        lines = list(inv.posted_gl_header.lines.order_by("line_no"))
        self.assertEqual(len(lines), 3)
        self.assertEqual(lines[2].account_id, self.coa_igst.account_code)
        self.assertEqual(lines[2].amount, Decimal("-180.00"))

    def test_bulk_post_two_invoices_sequential_tran_id(self):
        a = self._gst_invoice(client=self.client_row, invoice_no="FYGL-A")
        b = self._gst_invoice(client=self.client_row, invoice_no="FYGL-B")
        n, errs = bulk_post_fresh_invoices_to_gl(invoice_pks=[b.pk, a.pk], user=self.user)
        self.assertEqual(n, 2)
        self.assertEqual(errs, [])
        a.refresh_from_db()
        b.refresh_from_db()
        self.assertEqual(a.posted_gl_header.tran_id, "Sales-1")
        self.assertEqual(b.posted_gl_header.tran_id, "Sales-2")

    def test_delete_posted_invoice_blocked(self):
        inv = self._gst_invoice(client=self.client_row, invoice_no="FYGL-DEL")
        bulk_post_fresh_invoices_to_gl(invoice_pks=[inv.pk], user=self.user)
        c = HttpClient()
        c.force_login(self.user)
        r = c.post(
            reverse("invoices"),
            {"action": "delete", "pk": str(inv.pk)},
        )
        self.assertEqual(r.status_code, 302)
        self.assertTrue(Invoice.objects.filter(pk=inv.pk).exists())


class PdfExportBatchTests(SimpleTestCase):
    @patch("playwright.sync_api.sync_playwright")
    def test_invoice_html_list_reuses_one_browser(self, mock_sync):
        mock_page = MagicMock()
        mock_page.pdf.return_value = b"%PDF-1.4"
        mock_browser = MagicMock()
        mock_browser.new_page.return_value = mock_page
        mock_p = MagicMock()
        mock_p.chromium.launch.return_value = mock_browser
        mgr = MagicMock()
        mgr.__enter__.return_value = mock_p
        mgr.__exit__.return_value = None
        mock_sync.return_value = mgr

        from sales.invoices.pdf_export import invoice_html_list_to_pdf_bytes

        out = invoice_html_list_to_pdf_bytes(html_documents=["<html>a</html>", "<html>b</html>"])
        self.assertEqual(out, [b"%PDF-1.4", b"%PDF-1.4"])
        self.assertEqual(mock_browser.new_page.call_count, 2)
        self.assertEqual(mock_page.close.call_count, 2)
        mock_browser.close.assert_called_once()
