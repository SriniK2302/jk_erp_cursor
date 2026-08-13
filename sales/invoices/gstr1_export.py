"""Download exports for the GSTR1-style invoice list report (CSV, TSV, plain text, Excel)."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO

from django.http import HttpResponse

HEADERS = [
    "Client GSTN",
    "Client name",
    "Inv No",
    "Inv Date",
    "Narration",
    "Taxable value",
    "CGST",
    "SGST",
    "IGST",
    "Inv gross",
]

EXPORT_FORMATS = frozenset({"csv", "tsv", "txt", "xlsx"})


def _money_str(v) -> str:
    if isinstance(v, Decimal):
        return format(v, "f")
    return str(v)


def _row_list(row: dict) -> list:
    d = row["invoice_date"]
    date_s = d.isoformat() if hasattr(d, "isoformat") else str(d)
    return [
        row.get("client_gstn") or "",
        row.get("client_name") or "",
        row.get("invoice_no") or "",
        date_s,
        (row.get("narration") or "").replace("\r\n", "\n").replace("\r", "\n"),
        _money_str(row["taxable_value"]),
        _money_str(row["cgst"]),
        _money_str(row["sgst"]),
        _money_str(row["igst"]),
        _money_str(row["inv_gross"]),
    ]


def _invoice_date_value(row: dict) -> date:
    """Normalize invoice_date to ``datetime.date`` for Excel / typing."""
    d = row["invoice_date"]
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if hasattr(d, "isoformat"):
        return date.fromisoformat(str(d)[:10])
    raise TypeError(f"Unsupported invoice_date type: {type(d)!r}")


def _excel_float(v) -> float:
    """Money / numeric fields as Python float for openpyxl (stored as Excel numbers)."""
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    return float(v)


def _row_for_xlsx(row: dict) -> list:
    """Row values for .xlsx: text columns as str, date as ``date``, amounts as float."""
    return [
        row.get("client_gstn") or "",
        row.get("client_name") or "",
        row.get("invoice_no") or "",
        _invoice_date_value(row),
        (row.get("narration") or "").replace("\r\n", "\n").replace("\r", "\n"),
        _excel_float(row.get("taxable_value")),
        _excel_float(row.get("cgst")),
        _excel_float(row.get("sgst")),
        _excel_float(row.get("igst")),
        _excel_float(row.get("inv_gross")),
    ]


# 1-based column indices for amount columns (F–J) and invoice date (D).
_XLSX_DATE_COL = 4
_XLSX_AMOUNT_COLS = (6, 7, 8, 9, 10)
_XLSX_NUM_FMT_MONEY = "#,##0.00"
_XLSX_NUM_FMT_DATE = "yyyy-mm-dd"


def safe_filename_base(name: str) -> str:
    return re.sub(r"[^\w.\-]+", "_", name).strip("._") or "gstr1_export"


def gstr1_export_http_response(
    *,
    report_rows: list[dict],
    export_fmt: str,
    filename_base: str,
) -> HttpResponse:
    """Build an ``HttpResponse`` attachment for the given rows and format."""
    fmt = (export_fmt or "").strip().lower()
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Unsupported export format: {export_fmt!r}")
    base = safe_filename_base(filename_base)

    if fmt == "csv":
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(HEADERS)
        for row in report_rows:
            writer.writerow(_row_list(row))
        payload = ("\ufeff" + buf.getvalue()).encode("utf-8")
        resp = HttpResponse(payload, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{base}.csv"'
        return resp

    if fmt == "tsv":
        buf = StringIO()
        writer = csv.writer(buf, delimiter="\t", lineterminator="\n")
        writer.writerow(HEADERS)
        for row in report_rows:
            writer.writerow(_row_list(row))
        resp = HttpResponse(buf.getvalue().encode("utf-8"), content_type="text/tab-separated-values; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{base}.tsv"'
        return resp

    if fmt == "txt":
        lines = ["\t".join(HEADERS)]
        for row in report_rows:
            lines.append("\t".join(_row_list(row)))
        body = "\n".join(lines) + "\n"
        resp = HttpResponse(body.encode("utf-8"), content_type="text/plain; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{base}.txt"'
        return resp

    # xlsx — use native date / float so Excel treats columns as dates and numbers, not text.
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "GSTR1"
    ws.append(HEADERS)
    for row in report_rows:
        ws.append(_row_for_xlsx(row))
    for excel_row in range(2, ws.max_row + 1):
        ws.cell(row=excel_row, column=_XLSX_DATE_COL).number_format = _XLSX_NUM_FMT_DATE
        for col in _XLSX_AMOUNT_COLS:
            ws.cell(row=excel_row, column=col).number_format = _XLSX_NUM_FMT_MONEY
    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{base}.xlsx"'
    return resp
